# -*- coding: GBK -*-

import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from BaseScraper import BaseScraper



class JDScraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, max_items=100):
        #super().__init__(headless=None, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.logger = self._init_logger()
        self.driver = self.init_driver()
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self.insert_image = True
        self._setup_excel()
        #self.lock = threading.Lock()  # 用于同步的锁

    def human_sleep(self, min_time=0.5, max_time=1.5):
        time.sleep(random.uniform(min_time, max_time))

    def wait_manual_verification(self, min_seconds=2, max_seconds=3):
        """
        检查当前页面是否跳转到京东的风控滑块验证页（risk_handler），
        如果是，则暂停程序等待用户手动滑动验证，通过后按回车继续。
        """
        if 'risk_handler' in self.driver.current_url:
            print(" 检测到滑块风控页面，已暂停程序，建议降低抓取速度，请手动完成滑块验证。")
            print(" 完成验证后，按下回车键继续...")
            input(" 等待中...")
            # 验证通过后，判断是否跳转回正常页面
            if 'risk_handler' in self.driver.current_url:
                print(" 仍处于风控页面，验证可能失败")
                return False
            else:
                print(" 验证通过，继续执行")
                return True
        else:
            self.human_sleep(min_seconds, max_seconds)
            return True



    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def save_cookies(self, path="JD_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)


    def load_cookies(self, path="JD_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):#没做
        print("正在尝试登录JD...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)
            input("如出现滑块，请手动完成验证后按 Enter 继续...")
            self.save_cookies()
        except Exception as e:
            self.logger.warning(print("登录失败:", e))

    def search(self):
        try:
            print("尝试用定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'key')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'button')))
        except Exception as e:
            print("搜索框定位失败")
        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.wait_manual_verification()
        except Exception as e:
            self.logger.warning(print("搜索操作失败:", e))

    def go_to_page(self, page_number):

        self.wait_manual_verification(1,2)
        try:
            # 定位到输入框并输入页数
            search_input = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//span[@class="p-skip"]//input[@class="input-txt"]')
            ))
            search_input.clear()  # 清空输入框
            search_input.send_keys(page_number)  

            # 定位到确认按钮并点击
            confirm_button = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//span[@class="p-skip"]//a[@class="btn btn-default"]')
            ))
            confirm_button.click()
            # 模拟人工休眠
            self.wait_manual_verification(1,2)
        except Exception as e:
            # 记录错误日志
            self.logger.warning(f"翻页失败: {e}")
            self.save_page_html("error_page.html")



    def get_more (self, url):
        post_text = ''
        main_window = self.driver.current_window_handle

        try:
            # 提取商品ID，防止href匹配失败
            item_id = re.search(r'/(\d+)\.html', url)
            if not item_id:
                self.logger.warning(print("无法提取商品ID，跳过"))
                return 0
            item_id = item_id.group(1)

            # 用模糊匹配定位链接
            link_xpath = f'//a[contains(@href, "{item_id}")]'
            link_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, link_xpath)))

            # Ctrl+点击打开新标签页
            ActionChains(self.driver).key_down(Keys.CONTROL).click(link_element).key_up(Keys.CONTROL).perform()

            self.wait_manual_verification(1,2)


            # 等待新标签页打开
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning(print("新标签页未打开"))
                self.save_page_html("error_page.html")

                return 0

            # 重新获取所有窗口句柄
            handles = self.driver.window_handles
            new_tabs = [h for h in handles if h != main_window]
            if not new_tabs:
                self.logger.warning(print("找不到新标签页句柄"))
                return 0

            new_tab = new_tabs[-1]
            self.driver.switch_to.window(new_tab)

            self.wait_manual_verification(2,3)


            # 等待页面刷新
            try:
                # 等待包邮标签出现
                self.wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//span[contains(@class, "free-shipping")]')
                ))

                # 获取包邮文字
                post_text = self.driver.find_element(
                    By.XPATH,
                    '//span[contains(@class, "free-shipping")]'
                ).text.strip()

            except Exception:
                post_text = ''

            return post_text  # 例如返回：'包邮'

        except Exception as e:
            self.logger.warning(print("获取包邮失败"))

        finally:
            # 关闭新标签页，切换回主窗口
            self.human_sleep(1,2)
            try:
                if self.driver.current_window_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    self.wait_manual_verification(1,2)
            except Exception as e:
                self.logger.warning(print("关闭标签页或切换主窗口失败"))
                self.save_page_html("error_page.html")
                

        return post_text

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.taobao.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(print(f"图片请求失败: {url}"))
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(print(f"图片解码失败（格式可能不受支持，如webp）: {url}"))
                return None
            except Exception as e:
                self.logger.error(print(f"图片处理异常: {e} | URL: {url}"))
                return None

            # 创建文件夹
            folder = "images"
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 生成文件名，保存为 jpg
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(print(f"图片下载异常: {e} | URL: {url}"))
            return None

    def clean_image_folder(self):
        folder = "images"
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # 删除文件或快捷方式
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # 删除子文件夹

    ##滑动窗口
    def scroll_step_down(self, base_step=300):
        """模拟人类向下小段滑动"""
        step = random.randint(base_step - 100, base_step + 100)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.6, 1.2))
        self.wait_manual_verification()


    def smart_scroll_until_loaded(self, min_new_items=4, max_scroll_attempts=10):
        """边滚动边检测新商品是否加载，最多尝试 max_scroll_attempts 次"""
        last_count = 0
        attempts = 0

    def multi_scroll_up_down(self, up_num=1,down_num=8 , up_step=-200, down_step=800):
        # 向下滚动
        for _ in range(down_num):
            self.scroll_step_down(down_step)
    
        # 向上滚动
        for _ in range(up_num):
            self.scroll_step_down(up_step)
 

        # 模拟人工休眠，防止过快的滚动
        time.sleep(random.uniform(1.0, 1.5))


    def parse_page(self, page_number):
        html = self.driver.page_source
        doc = pq(html)
        items = doc('ul.gl-warp.clearfix > li').items()
        slide_counter = 0  # 每页开始前初始化下滑计数器

        for item in items:
            if self.count - 2 >= self.max_items:
                print(f"已达到最大抓取数量：{self.max_items}，停止抓取。")
                return False

#            if item.find('.title--RoseSo8H').text() or item.find('.headTitleText--hxVemljn').text():
#                continue

            # 获取图片标签
            img_tag = item.find('img')

            # 处理懒加载情况，优先选择 data-lazy-img 或者 src
            img_url = img_tag.attr('data-lazy-img') if img_tag.attr('data-lazy-img') else img_tag.attr('src')

            # 判断是否为相对路径，如果是，补全为完整的 URL
            if img_url and img_url.startswith('//'):
                img_url = 'https:' + img_url  # 补全相对路径为完整 URL

            # 打印或保存图片链接
            print(img_url)

            # 依次尝试 src / data-src / src2 / data-ks-lazyload
            img_url = (
                img_tag.attr('src')
                or img_tag.attr('data-src')
                or img_tag.attr('src2')
                or img_tag.attr('data-ks-lazyload')
                or img_tag.attr('data-lazyload')
                or ""
            ).strip()

            # 修复以 // 开头或缺少协议的情况
            if img_url.startswith("//"):
                img_url = "https:" + img_url
            elif img_url.startswith("/"):
                img_url = "https://img.alicdn.com" + img_url

            # 判断图片链接是否有效并下载到本地
            if not img_url or not isinstance(img_url, str) or not img_url.startswith(('/', '//', 'http')):
                self.logger.warning(print(f"[警告] 第 {self.count - 1} 个商品图片 URL 无效，跳过插图"))
                self.logger.warning(print("商品 HTML 片段："))
                print(item.outer_html())
                image_path = None
            else:
                image_path = self.download_image(img_url, self.count - 1) if self.insert_image else None

            #标题
            title_elem = item.find('.p-name em')
            title = title_elem.text().strip() if title_elem else ''
            #价格
            price_elem = item.find('.p-price i')
            price_text = price_elem.text().strip() if price_elem else ''          
            try:
                price = float(price_text)
            except ValueError:
                price = 0.0

            #京东不显示销量
            deal = 0
            #无地址
            location = ''
            #店铺名称和链接
            shop_elem = item.find('.p-shop a')
            shop_name = shop_elem.text().strip() if shop_elem else ''
            shop = shop_name
            shop_href = shop_elem.attr('href') if shop_elem else ''
            shop_url = 'https:' + shop_href if shop_href and shop_href.startswith('//') else shop_href
            #评论数量
            num_com = item.find('.p-commit a')
            num_com = num_com.text().strip() if num_com else ''
            #商品链接
            item_url = item.find('a[href^="//item.jd.com/"]')
            item_url = item_url.attr('href') if item_url else ''
            item_url = 'https:' + item_url if item_url and item_url.startswith('//') else item_url

            post = self.get_more(item_url)


            row = self.count
            self.sheet.row_dimensions[row].height = 65
            self.sheet.column_dimensions['L'].width = 13

            self.sheet.cell(row=row, column=1, value=row - 1)
            self.sheet.cell(row=row, column=2, value=title)
            self.sheet.cell(row=row, column=3, value=price)
            self.sheet.cell(row=row, column=4, value=deal)
            self.sheet.cell(row=row, column=5, value=location)
            self.sheet.cell(row=row, column=6, value=shop)
            self.sheet.cell(row=row, column=7, value=post)
            self.sheet.cell(row=row, column=8, value=item_url)
            self.sheet.cell(row=row, column=9, value=shop_url)
            self.sheet.cell(row=row, column=10, value=img_url)
            self.sheet.cell(row=row, column=11, value=num_com)

            if self.insert_image and image_path and os.path.exists(image_path):
                try:
                    img = ExcelImage(image_path)
                    img.width, img.height = 80, 80
                    self.sheet.add_image(img, f'L{row}')
                except Exception as e:
                    self.logger.error(print(f"插入图片失败: {e}（图片路径: {image_path}）"))

            print(f"第{row - 1}个商品信息已保存")
            self.count += 1
            slide_counter += 1  # 累加计数
            if slide_counter % 6 == 0:
                #print(f"已抓取 {slide_counter} 个商品，调用一次下滑函数...")
                self.scroll_step_down()
            self.human_sleep(1, 2)

        return True

    def turn_page(self, page_number):
        try:
            # 找到页码输入框
            input_box = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#J_bottomPage .p-skip .input-txt')
            ))
            input_box.clear()
            input_box.send_keys(str(page_number))

            # 点击“确定”按钮
            confirm_btn = self.driver.find_element(
                By.CSS_SELECTOR, '#J_bottomPage .p-skip .btn'
            )
            confirm_btn.click()
            self.wait_manual_verification(1,2)

            # 等待当前页码变更为目标页
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, '#J_bottomPage .p-num .curr'), str(page_number)
            ))

            # 停顿防止频繁操作
            self.human_sleep(2, 3)

        except Exception as e:
            self.logger.error(f"翻页失败，目标页码 {page_number}：{e}")
            self.save_page_html("error_page.html")


    def run(self):
        self.driver.get("https://www.jd.com/")
        input("请在浏览器中手动登录京东，如出现滑块，请手动完成验证后按 Enter 继续......")
        self.search()

        if self.page_start != 1:
            self.multi_scroll_up_down()
            self.go_to_page(self.page_start)

        for page in range(self.page_start, self.page_end + 1):
            if self.count - 2 >= self.max_items:  # 检查是否已达到最大抓取数量
                print(f"已抓取 {self.max_items} 个商品，停止抓取")
                break  # 达到最大数量时停止抓取

            self.multi_scroll_up_down(8,8,-800,800)
            proceed = self.parse_page(page)
            if not proceed:
                self.logger.warning(print(f"第 {page} 页爬取失败，跳过"))
                continue  # 如果该页抓取失败，跳过

            if page != self.page_end:
                self.turn_page(page + 1)

        filename = f"{self.keyword}_{time.strftime('%Y%m%d-%H%M')}.xlsx"
        self.excel.save(filename)
        print(f"文件已保存: {filename}")
        self.clean_image_folder()
        self.driver.quit()


if __name__ == "__main__":
    keyword = "奉贤黄桃"#input("输入搜索关键词：")
    start_page =1 # int(input("起始页码："))
    end_page = 3 #int(input("终止页码："))
    max_items = 100 #int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"

    scraper = JDScraper(keyword, start_page, end_page, max_items=max_items)
    scraper.insert_image = insert_image
    scraper.run()

