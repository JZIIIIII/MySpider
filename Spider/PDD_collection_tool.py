# -*- coding: utf-8 -*-
import time
import random
import re
import requests
import os
import shutil
import regex
import json
from datetime import datetime
from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError
from tkinter import filedialog
from fuzzywuzzy import fuzz

from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from pyquery import PyQuery as pq
from openpyxl import Workbook
from Spider.BaseScraper  import BaseScraper
from Spider.AntiScrapingException import CaptchaHandler
from path_utils import static_image_path ,static_hash_path , static_excel_path


class PDDScraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, max_items=100, insert_image=True, pierce_span=[0,0]):
        super().__init__(headless=True, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.pierce_span = pierce_span
        self.insert_image = insert_image
        self.count = 2
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self._setup_excel()
        self.empty_data_count = 0
        self.anti_spider_triggered = False
        self.captcha_handler = CaptchaHandler(self.driver, self.logger)
        self.captcha_handler.PDDsliderl()  # 调用风控

    # 定义 XPATH 常量
    XPATH_SALES_PATTERNS = [
        '//div[contains(@class,"AsbGpQv_")]/span[contains(text(),"已拼")]',
        '//div[contains(@class,"AsbGpQv_")]/span[contains(text(),"已抢")]',
        '//div[contains(@class,"BD_8SBr6")]/span[contains(text(),"总售")]',
        '//div[contains(@class,"BD_8SBr6")]/span[2][contains(text(),"已抢") or contains(text(),"已拼")]',
        '//div[contains(@class,"BD_8SBr6")]/span[contains(text(),"已拼")]',
    ]
    
    def _setup_excel(self):
        # 根据修改后的 save_to_excel 调整表头顺序
        headers = ['Num', 'Title', 'Price', 'Deal', 'shop_url', 'CommentNum', 'ShopName', 'Postage', 'Tags', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def simulate_click(self, element):
        try:
            actions = ActionChains(self.driver)
            actions.move_to_element(element).click().perform()
            self.logger.info("模拟鼠标点击成功")
            return True
        except Exception as e:
            self.logger.error(f"模拟鼠标点击失败:{e}")
            self.save_page_html("error_page.html")

            return False

    def click_fake_search_box(self, timeout=10):
        wait = WebDriverWait(self.driver, timeout)
        try:
            time.sleep(2)  # 等待页面加载

            # 等待目标元素可点击
            element = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "_18v23kPu")))

            # 调用封装好的模拟点击函数
            return self.simulate_click(element)
            self.RiskPause(self.captcha_handler.PDDsliderl())
        except Exception as e:
            self.logger.error(f"搜索框模拟点击失败:{e}")
            self.save_page_html("error_page.html")

            return False

    def scroll_step_down(self, base_step=800):
        """模拟人类向下较大幅度滑动"""
        step = random.randint(base_step - 300, base_step + 300)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.8, 1.5))  # 适当增加等待时间，保证加载

    def search(self):
        try:
            # 先尝试定位真实搜索框
            self.logger.info("尝试定位真实搜索框...")
            real_search_box = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@type='search' and contains(@class, '_2bfwu6WT')]")
            ))
        except Exception:

            self.logger.info("请尝试点击首页搜索框进入搜索页面...")
            return  # 找不到搜索框，结束函数，等待用户点击假搜索框进入搜索页

        try:
            # 定位搜索按钮
            search_btn = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'RuSDrtii') and text()='搜索']")
            ))
            # 输入关键词
            self.logger.info(f"输入关键词: {self.keyword}")
            real_search_box.clear()
            real_search_box.send_keys(self.keyword)

            # 等待 1~2 秒，模拟人操作
            self.human_sleep(1, 2)

            # 用你的 simulate_click 函数点击搜索按钮
            self.simulate_click(search_btn)

            # 等待 1~2 秒，等待搜索结果加载
            self.human_sleep(1, 2)

            self.logger.info("搜索成功")

        except Exception as e:
            self.logger.error(f"搜索失败:{e}")
            self.save_page_html("error_page.html")

    def inject_pdd_fetch_xhr_interceptor(self, driver, max_cache=50):
        """
        注入拦截器，捕获 fetch 和 XMLHttpRequest 类型的 /proxy/api/search 响应。
        每次响应缓存到数组中，最多保留 max_cache 条，最新的在数组末尾。
        """
        driver.execute_script(f"""
        (function() {{
            if (window._pdd_interceptor_installed) return;
            window._pdd_interceptor_installed = true;

            // 缓存数组
            window._pdd_responses = [];

            // ====== fetch 拦截 ======
            const origFetch = window.fetch;
            window.fetch = async (...args) => {{
                const response = await origFetch(...args);
                try {{
                    const url = args[0];
                    if (typeof url === 'string' && url.includes('/proxy/api/search')) {{
                        const cloned = response.clone();
                        cloned.json().then(data => {{
                            if (data) {{
                                if (window._pdd_responses.length >= {max_cache}) window._pdd_responses.shift();
                                window._pdd_responses.push(data);
                                console.log('[拦截到 fetch 搜索响应]', data);
                            }}
                        }}).catch(err => console.warn('[fetch JSON 解析失败]', err));
                    }}
                }} catch (err) {{ console.warn('[fetch 拦截异常]', err); }}
                return response;
            }};

            // ====== XHR 拦截 ======
            const origXHRSend = XMLHttpRequest.prototype.send;
            XMLHttpRequest.prototype.send = function(body) {{
                this.addEventListener('load', function() {{
                    try {{
                        const url = this.responseURL || '';
                        if (url.includes('/proxy/api/search') && this.responseType === '' && this.responseText) {{
                            const data = JSON.parse(this.responseText);
                            if (data) {{
                                if (window._pdd_responses.length >= {max_cache}) window._pdd_responses.shift();
                                window._pdd_responses.push(data);
                                console.log('[拦截到 XHR 搜索响应]', data);
                            }}
                        }}
                    }} catch(e) {{ console.warn('[XHR 解析失败]', e); }}
                }});
                origXHRSend.apply(this, arguments);
            }};
        }})();
        """)

    def get_pdd_search_responses(self, driver):
        """返回缓存的响应数组（可能为空），最新的在数组末尾。"""
        return driver.execute_script("return window._pdd_responses || [];")

    def clear_pdd_search_responses(self, driver):
        """清空缓存，避免重用旧数据。"""
        driver.execute_script("window._pdd_responses = [];")

    def dump_pdd_item(self, g, max_dump=3):
        """
        调试用：保存 PDD 原始 item 到 JSON 文件
        """
        if not g:
            return

        import json, os
        from datetime import datetime

        if not hasattr(self, "_pdd_dump_count"):
            self._pdd_dump_count = 0

        if self._pdd_dump_count >= max_dump:
            return

        os.makedirs("debug_dump", exist_ok=True)
        path = "debug_dump/pdd_raw_items.json"

        record = {
            "_dump_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "item": g
        }

        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                except:
                    data = []
        else:
            data = []

        data.append(record)

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        self._pdd_dump_count += 1

    def parse_all_showcases(self, max_items=100, platform='PDD', hash_json=None, max_empty_scrolls=5):
        """
        采集 PDD 商品信息，最多 max_items 条。
        支持 fetch/XHR 懒加载补充，处理滚动到底情况。
        max_empty_scrolls: 连续队列为空超过这个次数则认为到底
        """
        results = []
        seen_titles = set()
        processed_count = 0
        empty_scroll_count = 0  # 连续空队列计数

        self.inject_pdd_fetch_xhr_interceptor(self.driver)

        if hash_json:
            hash_json = static_hash_path(hash_json)
            hash_set = self.load_hash_set(hash_json)
        else:
            hash_set = set()

        while len(results) < max_items:
            # 风控/暂停检测
            self.RiskPause(self.captcha_handler.PDDsliderl())
            self.RiskPause(self.anti_spider_triggered)
            self.wait_if_paused()
            if self.should_stop():
                self.logger.info("检测到提前终止命令，保存已抓取内容并退出 parse_all_showcases。")
                if hash_json:
                    self.save_hash_set(hash_set, hash_json)
                break

            # 取下一条商品
            g = self.get_next_PDD_item(clear_after_fetch=True)

            #  调试：保存原始 PDD item 到 json（不打 log）
            #self.dump_pdd_item(g)

            queue_len = len(getattr(self, "_pdd_items_queue", []))
            self.logger.info(f"[队列状态] _pdd_items_queue长度: {queue_len}, 已采集: {len(results)}")

            if not g:
                empty_scroll_count += 1
                if empty_scroll_count >= max_empty_scrolls:
                    self.logger.info(f"连续 {max_empty_scrolls} 次滚动未获取新商品，认为已到底，停止采集。")
                    break
                # 队列空 -> 下滑加载更多
                self.scroll_step_down(base_step=1200)
                time.sleep(2)
                continue

            empty_scroll_count = 0  # 重置空队列计数

            title = g.get("title", "").strip()
            if not title or title in seen_titles:
                continue
            seen_titles.add(title)


            # ==== 价格区间过滤 ====
            min_p, max_p = self.pierce_span

            # 防止用户输入反区间 [100, 50]
            if min_p > max_p:
                min_p, max_p = max_p, min_p

            # 只有当区间不是 [0,0] 时才进行过滤
            if not (min_p == 0 and max_p == 0):
                try:
                    price_value = float(g.get("price", 0))
                except:
                    continue  # 无法解析价格 → 跳过

                # 不在闭区间 → 跳过该商品
                if not (min_p <= price_value <= max_p):
                    self.logger.info(f"[价格过滤] {price_value} 不在区间 [{min_p}, {max_p}] 内，跳过")
                    continue
            # ======================

            # 哈希去重
            item_dict = {'title': title, 'price': g.get("price", "")}
            features_url = g.get("img_url", "") or ""
            item_hash = self.compute_hash(item_dict, platform, features_url)
            if item_hash in hash_set:
                self.logger.info(f"[跳过] 已存在商品: {title}")
                continue
            hash_set.add(item_hash)

            # 下载图片
            img_url = g.get("img_url", "")
            img_path = None
            if self.insert_image and img_url:
                img_path = self.download_image(img_url, title)

            # 组装结果
            result = {
                'title': title,
                'price': g.get("price", ""),
                'deal_num': g.get("deal", 0),
                'shop_url': g.get("shop_url", ""),
                'comment_count': g.get("comment_count", ""),
                'shop_name': g.get("shop_name", ""),
                'postage_info': "包邮" if g.get("free_shipping") else "",
                'tags': "",
                'img_url': img_url,
                'img_path': img_path
            }
            results.append(result)
            processed_count += 1
            self.count = processed_count
            self.logger.info(f"[{processed_count}] 采集到商品: {title}")

            # 每采 2 条滚动一次
            if processed_count % 2 == 0:
                self.scroll_step_down(base_step=800)
                time.sleep(2)

        self.logger.info(f"\n共提取到 {len(results)} 个商品（上限：{max_items}）")
        if hash_json:
            self.save_hash_set(hash_set, hash_json)

        return results



    def save_to_excel(self, results, filename='results.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = '商品信息'

        # 表头
        headers = ['标题', '价格', '成交量', '店铺链接', '评论数量', '店铺名称', '包邮信息', '标签', '图片']
        ws.append(headers)

        # 设置图片插入列列宽（第9列）
        img_col_letter = 'I'
        ws.column_dimensions[img_col_letter].width = 15

        for i, item in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=item['title'])
            ws.cell(row=i, column=2, value=item['price'])
            ws.cell(row=i, column=3, value=item['deal_num'])
            ws.cell(row=i, column=4, value=str(item['shop_url']))
            ws.cell(row=i, column=5, value=item['comment_count'])
            ws.cell(row=i, column=6, value=item['shop_name'])
            ws.cell(row=i, column=7, value=item['postage_info'])
            ws.cell(row=i, column=8, value=item['tags'])

            # 如果开启插图功能
            if self.insert_image:
                img_path = item.get('img_path')
                if img_path and os.path.exists(img_path):
                    try:
                        # 尝试用 Pillow 打开并重新编码为 JPEG（修复部分 JPEG 插入失败问题）
                        with Image.open(img_path) as img:
                            rgb_img = img.convert("RGB")
                            rgb_img.save(img_path, format='JPEG')  # 覆盖原图
                    
                        # 插入图片
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 80
                        excel_img.height = 80
                        ws.row_dimensions[i].height = 60
                        excel_img.anchor = f'{img_col_letter}{i}'
                        ws.add_image(excel_img)
                    except Exception as e:
                        self.logger.error(f"[!] 图片插入失败: {img_path} | 错误: {e}")
                else:
                    self.logger.warning(f"[!] 图片路径无效或文件不存在: {img_path}")
            
        # 保存 Excel 文件
        wb.save(filename)

        self.logger.info(f"[√] 数据已保存到 Excel：{filename}")

    def webp_to_png(self, webp_bytes, save_path):
        try:
            img = Image.open(BytesIO(webp_bytes))
            if img.format != 'WEBP':
                self.logger.warning(f"警告：不是WEBP格式，实际是 {img.format}")
            img = img.convert("RGBA")  # 保留透明度
            img.save(save_path, format="PNG")
            self.logger.info(f"WEBP图片转换PNG并保存成功: {save_path}")
            return save_path
        except Exception as e:
            self.logger.warning(f"WEBP转PNG失败: {e}")
            return None

    def download_image(self, img_url, title, platform='PDD'):
        """
        下载图片并保存到 images/<platform>/ 目录下，文件名由标题生成。
        若为 WebP 格式将自动转为 PNG，其余转为 JPG。
        """
        # 清理标题中的特殊字符，避免路径错误
        filename = re.sub(r'[\\/:*?"<>|]', '_', title[:20].strip().replace(' ', '_')) + ".jpg"

        # 硬编码路径，直接指定保存路径
        folder = f"images/{platform}"  # 使用硬编码的路径
        if not os.path.exists(folder):
            os.makedirs(folder)  # 如果目录不存在，则创建

        path = os.path.join(folder, filename)  # 生成完整路径

        # 拼多多常见图片链接修复
        if img_url.startswith("//"):
            img_url = "https:" + img_url
        elif img_url.startswith("/"):
            img_url = "https://img.pddpic.com" + img_url

        img_url = img_url.split('?')[0]  # 去掉查询参数

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "Referer": "https://mobile.pinduoduo.com/",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        }

        try:
            resp = requests.get(img_url, headers=headers, timeout=10)
            if resp.status_code != 200:
                self.logger.error(f"图片下载失败，状态码: {resp.status_code} | URL: {img_url}")
                return None

            try:
                img = Image.open(BytesIO(resp.content))

                if img.format == 'WEBP':
                    self.logger.info(f"[!] 图片为WEBP，转换为PNG: {img_url}")
                    png_path = os.path.splitext(path)[0] + ".png"
                    result = self.webp_to_png(resp.content, png_path)
                    return result

                img = img.convert('RGB')
                img.save(path, format='JPEG')
                self.logger.info(f"图片已保存: {path}")
                return path

            except UnidentifiedImageError:
                self.logger.error(f"[!] 图片格式不支持或路径无效: {img_url}")
                return None
            except Exception as e:
                self.logger.error(f"[!] 图片处理异常: {e} | URL: {img_url}")
                return None

        except Exception as e:
            self.logger.error(f"[!] 图片下载异常: {e} | URL: {img_url}")
            return None

    def clean_url(self, url):
        """
        使用正则表达式清理 URL，保留 goods_id 参数，去掉其他多余的参数。
        """
        try:
            # 使用正则匹配并提取 base_url 和 goods_id 参数
            match = re.search(r"(https?://[^\?]+)(\?[^#]*)", url)
            if match:
                base_url = match.group(1)  # 基础 URL
                query_string = match.group(2)  # 查询部分

                # 解析查询参数
                query_params = parse_qs(query_string[1:])  # 去掉 '?'
                
                # 只保留 'goods_id' 参数
                cleaned_params = {key: value for key, value in query_params.items() if key == 'goods_id'}
                
                # 构造新的查询字符串
                cleaned_query = urlencode(cleaned_params, doseq=True)
                
                # 如果有有效的 cleaned_query，则重建新的 URL
                if cleaned_query:
                    cleaned_url = f"{base_url}?{cleaned_query}"
                else:
                    cleaned_url = base_url  # 如果没有有效的参数，返回没有参数的 URL
                
                return cleaned_url
            else:
                return url  # 如果没有匹配，返回原始 URL
        except Exception as e:
            self.logger.error(f"Error cleaning URL: {e}")
            return url

    def parse_pdd_search_response_data(self, search_data):
        """
        解析 fetch 拦截到的 PDD 搜索响应，返回商品原始信息。
        """
        results = []
        try:
            items = search_data.get("items") or []
            for item in items:
                item_data = item.get("item_data", {})
                goods_model = item_data.get("goods_model", {})

                results.append({
                    "goods_name": goods_model.get("goods_name", ""),
                    "link_url": goods_model.get("link_url", ""),
                    "mall_name": goods_model.get("mall_name", ""),
                    "sales_tip": goods_model.get("sales_tip", ""),
                    "sales": goods_model.get("sales", 0),
                    "price_info": goods_model.get("price_info", 0),
                    "hd_thumb_url": goods_model.get("hd_thumb_url", ""),
                    "hd_url": goods_model.get("hd_url", ""),
                    "tag_list": goods_model.get("tag_list", []),
                })
        except Exception as e:
            self.logger.warning(f"[!] 解析 PDD 搜索响应失败: {repr(e)}")
        return results

    def normalize_text(self, s):
        """统一商品名，便于匹配"""
        if not s:
            return ""
        s = s.lower().strip()
        # 去掉空格、连字符、下划线
        s = re.sub(r'[\s\-–_]+', '', s)
        # 去掉 Unicode 控制符、符号
        s = regex.sub(r'\p{C}|\p{So}', '', s)
        # 保留中文、字母、数字
        s = regex.sub(r'[^\p{IsHan}\w]', '', s)
        return s

    def fuzzy_match(self, a, b, threshold=80):
        """模糊匹配"""
        score = max(fuzz.partial_ratio(a, b), fuzz.token_set_ratio(a, b))
        return score >= threshold

    
    def get_more(self, element):
        """
        进入详情页，获取评论数、店铺名、包邮信息、店铺链接、销量信息。
        包含滚动点击、异常捕获、风控处理、页面返回等逻辑。
        返回：
            comment_count (int): 评论数
            shop_name (str): 店铺名称
            postage_info (str): 包邮信息
            shop_url (str): 店铺链接
            deal_text (str): 销量信息文本
        """
        comment_count = 0
        shop_name = ""
        postage_info = ""
        shop_url = ""
        deal_text = ""

        try:
            # 滚动元素至可视区中间，避免被遮挡
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.5)

            # 使用JS点击元素，绕过遮挡问题
            self.driver.execute_script("arguments[0].click();", element)

            # 获取当前页面URL，清理后保存为店铺链接
            current_url = self.driver.current_url
            shop_url = self.clean_url(current_url)

            # 风控处理及暂停检测
            self.RiskPause(self.captcha_handler.PDDsliderl())
            self.wait_if_paused()

            if self.should_stop():
                self.logger.info("用户选择终止爬虫，提前结束运行。")
                self.stop()
                return 0, None, None, None, ""

            # 等待详情页关键元素加载，避免页面未完全加载导致找不到元素
            # 这里用不同选择器分步等待，捕获异常时记录日志并使用默认值

            # 获取销量信息（支持多个xpath尝试）
            try:
                for xp in self.XPATH_SALES_PATTERNS:
                    try:
                        elem = self.driver.find_element(By.XPATH, xp)
                        deal_text = elem.text.strip()
                        if deal_text:
                            break
                    except Exception:
                        continue
                if not deal_text:
                    deal_text = "销量信息未找到"
            except Exception as e:
                self.logger.warning(f"[!] 获取销量信息失败: {repr(e)}")
                deal_text = "销量信息异常"

            self.logger.info(f"获取销量: {deal_text}")

            # 获取评论数
            try:
                comment_text_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.F2MXl7Xc'))
                )
                comment_text = comment_text_element.text
                match = re.search(r'\((\d+)\)', comment_text)
                if match:
                    comment_count = int(match.group(1))
                else:
                    comment_count = 0
            except Exception:
                self.logger.warning("[!] 无法获取评论数，默认设置为0")
                comment_count = 0

            # 获取店铺名称
            try:
                shop_name_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.BAq4Lzv7'))
                )
                shop_name = shop_name_element.text.strip()
            except Exception as e:
                self.logger.warning(f"[!] 获取店铺名称失败: {repr(e)}")
                shop_name = "无店铺名称"

            # 获取包邮信息
            try:
                postage_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.RbQ7MTuU span.KDFIGUNK'))
                )
                postage_info = ' '.join([e.text.strip() for e in postage_elements if e.text.strip()])
                if not postage_info:
                    postage_info = "无包邮信息"
            except Exception as e:
                self.logger.warning(f"[!] 获取包邮信息失败: {repr(e)}")
                postage_info = "无包邮信息"

        except Exception as e:
            self.logger.error(f"[!] 获取详情数据失败: {repr(e)}")
            self.save_page_html("error_page.html")

        finally:
            # 返回列表页，等待主页面加载
            self.driver.back()
            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.rjNMXsUm._1unt3Js-')))
            except TimeoutException:
                time.sleep(1)
            self.human_sleep(1, 2)

        # 检测连续多次获取空数据，触发反爬逻辑
        if (comment_count == 0 and shop_name in ["", "无店铺名称"] and
            postage_info in ["", "无包邮信息"] and shop_url == ""):
            self.empty_data_count = getattr(self, 'empty_data_count', 0) + 1
            self.logger.warning(f"[!] 获取为空数据 {self.empty_data_count}/3 次")
            if self.empty_data_count >= 3:
                self.logger.error("[!] 连续 3 次获取为空，程序即将暂停")
                self.anti_spider_triggered = True
                return 0, None, None, None, ""
        else:
            # 有效数据则重置计数
            self.empty_data_count = 0

        return comment_count, shop_name, postage_info, shop_url, deal_text


    def get_next_PDD_item(self, clear_after_fetch=True):
        """
        返回下一条商品信息: title, img_url, shop_url, deal, price, free_shipping
        并在日志中打印队列状态
        """
        # 初始化队列
        if not hasattr(self, "_pdd_items_queue"):
            self._pdd_items_queue = []

            # 首次从 draw 初始化前20条
            draw_items = self.driver.execute_script(
                "return window.rawData?.stores?.store?.data?.ssrListData?.list || [];"
            ) or []

            for item in draw_items:
                obj = {
                    "title": item.get("goodsName") or "",
                    "img_url": item.get("imgUrl") or item.get("hd_url") or "",
                    "shop_url": "https://mobile.pinduoduo.com/" + (item.get("linkURL") or ""),
                    "deal": item.get("salesTip") or item.get("sales") or 0,
                    "price": item.get("priceInfo") or 0,
                    "free_shipping": any("包邮" in tag.get("text", "") for tag in item.get("tagList", []))
                }
                self._pdd_items_queue.append(obj)

            self.logger.info(f"[PDD] 队列初始化完成，draw 获取 {len(draw_items)} 条商品，总队列长度: {len(self._pdd_items_queue)}")

        # 从 fetch/XHR 补充新 items
        fetch_items = []
        responses = self.get_pdd_search_responses(self.driver)
        for search_data in reversed(responses):
            items = self.parse_pdd_search_response_data(search_data)
            '''
            os.makedirs("debug_dump", exist_ok=True)
            with open("debug_dump/pdd_parsed_items.json", "w", encoding="utf-8") as f:
                json.dump(items[:3], f, ensure_ascii=False, indent=2)
            '''
            fetch_items.extend(items)

        if fetch_items:
            self._pdd_items_queue.extend([
                {
                    "title": g.get("goods_name", ""),
                    "img_url": g.get("hd_thumb_url") or g.get("hd_url") or "",
                    "shop_url": g.get("link_url", ""),
                    "deal": g.get("sales_tip") or g.get("sales", 0),
                    "price": g.get("price_info", 0),
                    "free_shipping": any("包邮" in tag.get("text", "") for tag in g.get("tag_list", []))
                                    if isinstance(g.get("tag_list", []), list) else False
                }
                for g in fetch_items
            ])
            self.logger.info(f"[PDD] fetch 补充 {len(fetch_items)} 条商品，总队列长度: {len(self._pdd_items_queue)}")

            if clear_after_fetch:
                self.clear_pdd_search_responses(self.driver)

        # 返回队列中的下一条商品
        if self._pdd_items_queue:
            g = self._pdd_items_queue.pop(0)
            if g["shop_url"] and not g["shop_url"].startswith("http"):
                g["shop_url"] = "https://mobile.pinduoduo.com/" + g["shop_url"].lstrip("/")
            self.logger.debug(f"[PDD] 返回下一条商品, 队列剩余长度: {len(self._pdd_items_queue)}")
            return g

        # 队列为空
        self.logger.debug("[PDD] 队列为空，没有商品可返回")
        return None

    def clear_image_cache(self, subfolder='PDD'):
        """
        清空指定平台的图片缓存目录，默认是 images/PDD
        """
        # 使用硬编码的路径，不再调用 static_image_path
        folder = f"images/{subfolder}"

        if os.path.exists(folder) and os.path.isdir(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    self.logger.warning(f"[!] 删除缓存文件失败: {file_path} | 错误: {e}")
        else:
            self.logger.error(f"[!] 文件夹不存在或不是目录：{folder}")


    

    def run(self):
        try:
            self.driver.get("https://mobile.pinduoduo.com/")
            self.logger.info("登录拼多多成功")
        except:
            self.logger.error("登录拼多多失败")

        self.wait_for_login()
        #input("IMPORT")

        self.click_fake_search_box()
        self.search()
        self.RiskPause(self.captcha_handler.PDDsliderl())
        data = self.parse_all_showcases(max_items=self.max_items)

        if data:
            # 使用硬编码路径
            filename = f"{self.keyword}_{time.strftime('%Y%m%d_%H%M')}.xlsx"
            folder = "excel"  # 硬编码路径
            if not os.path.exists(folder):
                os.makedirs(folder)  # 如果目录不存在，则创建

            # 生成完整的文件路径
            filepath = os.path.join(folder, filename)

            self.save_to_excel(data, filepath)  # 保存数据到 Excel
        self.clear_image_cache()  # 清理图片缓存
        self.driver.close()
        self.driver.quit()



if __name__ == "__main__":
    kw = input("输入关键词：")
    sp = 1 #int(input("起始页码："))
    ep = 1 #int(input("结束页码："))
    mi = int(input("最大商品数："))
    show_img = input("是否插入图片 (y/n)：").strip().lower() == 'y'
    pierce_span=[100,300]
    spider = PDDScraper(kw, sp, ep, mi, show_img,pierce_span=pierce_span)
    spider.run()

