# -*- coding: utf-8 -*-
# 代码说明：
'''
代码功能： 基于ChromeDriver爬取taobao（淘宝）平台商品列表数据
输入参数:  KEYWORLD --> 搜索商品“关键词”；
          pageStart --> 爬取起始页；
          pageEnd --> 爬取终止页；修改代码让图片获取的步骤在            deal = item.find('.realSales--XZJiepmt').text().replace("万", "0000").split("人")[0].split("+")[0]
            deal = int(deal) if deal.isdigit() else 0
            location = item.find('.procity--wlcT2xH9 span').text()
            shop = item.find('.shopNameText--DmtlsDKm').text()
            post = "包邮" if "包邮" in item.find('.subIconWrapper--Vl8zAdQn').text() else "/"
            item_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href')
            shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')
            img_url = item.find('.mainPicAdaptWrapper--V_ayd2hD img').attr('src')

            num_com = self.get_comment_count(item_url)此处实现
 
输出文件：爬取商品列表数据
        'Page'        ：页码
        'Num'         ：序号
        'title'       ：商品标题
        'Price'       ：商品价格
        'Deal'        ：商品销量
        'Location'    ：地理位置
        'Shop'        ：店铺
        'IsPostFree'  ：是否包邮
        'Title_URL'   ：商品详细页链接
        'Shop_URL'    ：商铺链接
        'Img_URL'     ：图片链接
        'Img'         ；图片
'''
from pickle import TRUE
import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from Spider.BaseScraper import BaseScraper
from selenium.common.exceptions import NoSuchElementException

from Spider.AntiScrapingException import CaptchaHandler
from path_utils import static_image_path ,static_hash_path , static_excel_path



class Bilibili(BaseScraper):
    def __init__(self, keyword, start_page, end_page,insert_image=True, max_items=100):
        super().__init__(headless=True, proxy=None)  # 调用父类初始化
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.insert_image = insert_image
        self.max_items = max_items
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self._setup_excel()
        self.hash_set = set()
        #self.anti_spider_triggered = False
        self.captcha_handler = CaptchaHandler(self.driver, self.logger)
        self.captcha_handler.Taosliderl()
        self.captcha_handler.TaoCheckRisk()




    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)




    def save_cookies(self, path="taobao_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="taobao_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):
        #print("正在尝试登录淘宝...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)

            input("如出现滑块，请手动完成验证后按 Enter 继续...")

            self.save_cookies()
        except Exception as e:
            self.logger.warning(f"登录失败: {e}")
            self.save_page_html("error_page.html")

    def search(self):
        try:
            self.logger.info(f"尝试用s方法定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'q')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="J_SearchForm"]/div/div[1]/button')))
        except Exception as e1:
            self.logger.warning(f"s方法失败: {e1}\n尝试用w方法定位...")
            try:
                search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'search_input')))
                search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'search-btn')))
            except Exception as e2:
                self.logger.warning(f"w方法也失败: {e2}")

                raise Exception("搜索框和按钮定位失败，搜索终止")

                self.save_page_html("error_page.html")

        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.human_sleep(2, 3)
        except Exception as e:
            self.logger.error(f"搜索操作失败:{e}")
            self.save_page_html("error_page.html")

    def go_to_page(self, page_number):
        try:
            # 等待跳页输入框出现
            page_input = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'span.next-input input')))
            page_input.clear()
            page_input.send_keys(str(page_number))
            self.human_sleep()

            # 找到并点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button.next-pagination-jump-go')))
            confirm_btn.click()

            # 等待当前页码变为目标页
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.next-pagination-display em'),
                str(page_number)))

            self.human_sleep(3, 4)

        except Exception as e:
            self.logger.error(f"翻页失败: {e}")
            self.save_page_html("error_page.html")
            
    def get_comment_count(self, item_id):
        """
        打开商品详情页并提取评论数。
        支持风控处理与暂停/终止控制。
        """
        count = 0
        deal = ''
        main_window = self.driver.current_window_handle

        try:
            # 定位包含 a 的 div
            card_xpath = f'//div[a[@data-spm-act-id="{item_id}"]]'
            card_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, card_xpath)))

            # 滚动到可见并点击
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card_element)
            actions = ActionChains(self.driver)
            actions.click(card_element).perform()

            self.human_sleep(1, 2)

            # 等待新标签页打开
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return 0

            # 切换到新标签页
            new_tab = [h for h in self.driver.window_handles if h != main_window][-1]
            self.driver.switch_to.window(new_tab)
            self.human_sleep(2, 3)

            # === 风控检查与等待控制 ===
            self.RiskPause(self.captcha_handler.Taosliderl())
            self.RiskPause(self.captcha_handler.TaoCheckRisk())

            self.wait_if_paused()
            if self.should_stop():
                self.logger.info("用户主动终止爬虫，跳过当前评论数解析")
                self.stop()
                return 0

            try:
                # 等待销量模块加载
                self.wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]')
                ))

                deal_text = self.driver.find_element(
                    By.XPATH,
                    '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]'
                ).text.strip()

                if not deal_text:
                    # 如果 text 为空，尝试获取 textContent
                    deal_text = self.driver.execute_script(
                        "return arguments[0].textContent;", 
                        self.driver.find_element(
                            By.XPATH,
                            '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]'
                        )
                    ).strip()

                self.logger.info(f"获取销量文本: {deal_text}")
                deal_text = deal_text.lstrip('·').strip()

            except Exception as e:
                self.logger.warning(f"获取销量失败: {e}")
                deal_text = "商品详情获取销量失败 可能为0"

            # 提取评论数
            try:
                # 等待评论模块加载
                self.wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//span[contains(text(),"有图") or contains(text(),"视频")]')
                ))

                comment_text = self.driver.find_element(
                    By.XPATH,
                    '//span[contains(text(),"有图") or contains(text(),"视频")]/preceding-sibling::span[1]'
                ).text

                # 解析评论数量
                m = re.search(r'\(([一-龥\d,.+万]+)\)', comment_text)
                if m:
                    num_str = m.group(1).replace('+', '').replace(',', '')
                    if '万' in num_str:
                        count = int(float(num_str.replace('万', '')) * 10000)
                    else:
                        count = int(num_str)
                else:
                    self.logger.info("未匹配到评论数文本，默认为 0")
                    count = 0

            except Exception as e:
                self.logger.warning(f"获取评论数失败: {e}")
                count = 0

        except Exception as e:
            self.logger.warning(f"获取详情失败：{e}")
            self.save_page_html("error_page.html")

        finally:
            # 关闭新标签页并返回主窗口
            self.human_sleep(1, 2)
            try:
                if self.driver.current_window_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
            except Exception as e:
                self.logger.error(f"关闭标签页或切换窗口失败：{e}")
                self.save_page_html("error_page.html")
        
        return deal_text, count

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.taobao.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(f"图片请求失败: {url}")
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(f"图片解码失败（格式可能不受支持，如webp）: {url}")
                return None
            except Exception as e:
                self.logger.error(f"图片处理异常: {e} | URL: {url}")
                return None

            # 使用硬编码路径创建文件夹
            folder = "images/TaoBao"  # 直接使用硬编码的路径
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 构建文件名和文件路径
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(f"图片下载异常: {e} | URL: {url}")
            return None

    def clean_image_folder(self):
        folder = "images/TaoBao"  # 直接使用硬编码的路径
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # 删除文件或快捷方式
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # 删除子文件夹


    def parse_page(self, page_number, platform='TaoBao',hash_json=None):
        html = self.driver.page_source
        doc = pq(html)
        items = doc('div.content--CUnfXXxv > div > div').items()
        #content--CUnfXXxv
        # 载入历史哈希
        if hash_json:
            hash_json = static_hash_path(hash_json)
            hash_set = self.load_hash_set(hash_json)
            self.hash_set.update(hash_set)  # 合并新的哈希池
        else:
            hash_set  = self.hash_set
        slide_counter = 0  # 每页开始前初始化下滑计数器

        for item in items:

            # 如果暂停，则等待恢复
            self.wait_if_paused()

            # 检查是否被强制终止
            if self.should_stop():
                self.logger.info("检测到提前终止命令，保存已抓取内容并退出 parse_page。")
                if hash_json:
                    self.save_hash_set(self.hash_set, hash_json)  
                return False

            try:
                if self.count - 2 >= self.max_items:

                    #print(f"已达到最大抓取数量：{self.max_items}，停止抓取。")
                    if hash_json:
                        self.save_hash_set(self.hash_set, hash_json)  
                    return False


                # 检查是否触发了强制风控
                '''
                if self.anti_spider_triggered == True:
                    print(f"触发强制风控账号冻结 应前往淘宝APP解封")
                    if hash_json:
                        self.save_hash_set(hash_set, hash_json)  
                    return False
                '''

                if item.find('.title--RoseSo8H').text() or item.find('.headTitleText--hxVemljn').text():
                    continue

                self.RiskPause(self.captcha_handler.Taosliderl())

                # 获取图片标签
                img_tag = item.find('.mainPicAdaptWrapper--V_ayd2hD img')
                if not img_tag:
                    img_tag = item.find('.imageSwitch--fJ9SrtEb img')
                if not img_tag:
                    img_tag = item.find('img')

                # 获取图片 URL，首先检查 src 或者 data-lazy-img
                img_url = (
                    img_tag.attr('src')  # img 标签的 src 属性
                    or img_tag.attr('data-lazy-img')  # img 标签的 data-lazy-img 属性（你的 HTML 示例中是这个）
                    or img_tag.attr('data-src')  # 其他可能的属性
                    or img_tag.attr('src2')
                    or img_tag.attr('data-ks-lazyload')
                    or img_tag.attr('data-lazyload')
                    or ""
                ).strip()

                # 修复以 // 开头或缺少协议的情况
                if img_url.startswith("//"):
                    img_url = "https:" + img_url
                elif img_url.startswith("/"):
                    img_url = "https://img.alicdn.com" + img_url

                title = item.find('.title--qJ7Xg_90 span').text()

                price_text = item.find('.innerPriceWrapper--aAJhHXD4').text().replace('\n', '')

                try:
                    price = float(price_text)
                except ValueError:
                    price = 0.0

                deal = ''


                location = item.find('.procity--wlcT2xH9 span').text()
                shop = item.find('.shopNameText--DmtlsDKm').text()
                post = "包邮" if "包邮" in item.find('.subIconWrapper--Vl8zAdQn').text() else "/"

                item_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href')
                a_tag = doc.find('.doubleCardWrapperAdapt--mEcC7olq')

                if a_tag:
                    item_id = a_tag.attr('data-spm-act-id')
                    self.logger.info(f"id == {item_id}") 
                else:
                    self.logger.info("未找到目标元素")

                #url补全
                if item_url.startswith("//"):
                    item_url = "https:" + item_url
                elif item_url.startswith("/"):
                    item_url = "https://item.taobao.com" + item_url

                shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')
                #url补全
                if shop_url.startswith("//"):
                    shop_url = "https:" + shop_url
                elif shop_url.startswith("/"):
                    shop_url = "https://click.simba.taobao.com" + shop_url

            
                # ==== 新增哈希去重 ====
                item_dict = {'title': title, 'price': price}
                features_url = shop_url or ''  # 用图片链接代替特征链接

                item_hash = self.compute_hash(item_dict, platform, features_url)
                if item_hash in hash_set:
                    # print(f"[跳过] 已存在的商品: {title}")
                    continue
                self.hash_set.add(item_hash)
                # ======================
                # 判断图片链接是否有效并下载到本地
                if not img_url or not isinstance(img_url, str) or not img_url.startswith(('/', '//', 'http')):
                    self.logger.warning(f"[警告] 第 {self.count - 1} 个商品图片 URL 无效，跳过插图")
                    self.logger.warning("商品 HTML 片段：")
                    #print(item.outer_html())
                    image_path = None
                else:
                    image_path = self.download_image(img_url, self.count - 1) if self.insert_image else None

                deal, num_com = self.get_comment_count(item_id)


                # 调用父类的保存方法
                self.save_product_to_excel(
                    row=self.count,
                    title=title,
                    price=price,
                    deal=deal,
                    location=location,
                    shop=shop,
                    post=post,
                    item_url=item_url,
                    shop_url=shop_url,
                    img_url=img_url,
                    num_com=num_com,
                    image_path=image_path
                )
                self.count += 1
                slide_counter += 1  # 累加计数
                if slide_counter % 6 == 0:
                    #print(f"已抓取 {slide_counter} 个商品，调用一次下滑函数...")
                    self.scroll_step_down()
                self.human_sleep(1, 2)
            except Exception as e:
                self.logger.warning(f"[!] 解析商品异常: {e}")
                continue

        # 保存哈希集，方便下次增量爬取
        if hash_json:
            self.save_hash_set(self.hash_set, hash_json)  # 使用类成员变量 self.hash_set

        return True

    def turn_page(self, page_number):
        try:
            # 找到“下一页”按钮（根据 class）
            next_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button.next-next:not([disabled])')))
        
            next_btn.click()

            # 等待当前页码显示变为指定页数
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.next-pagination-display em'),
                str(page_number)))

            self.human_sleep(2, 3)

        except Exception as e:
            self.logger.error(f"下一页点击失败: {e}")


    def run(self):
        self.driver.get("https://live.bilibili.com/24531864?live_from=81002&trackid=web_pegasus_2.router-web-pegasus-2187328-75f4ffdddf-5jkg7.1755427701308.795&spm_id_from=333.1007.tianma.6-3-19.click")
        # 等待前端登录完成
        input("请在浏览器中完成登录后，按回车继续...")

        like_btn = WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.like-btn"))
        )

        try:
            # 循环点击 1000 次
            for i in range(1000):
                actions = ActionChains(self.driver)
                (
                    actions.move_to_element(like_btn)             # 移动到按钮
                           .click(like_btn)                       # 点击按钮
                           .pause(random.uniform(0.001, 0.002))   # 暂停 2~3 毫秒
                )
                actions.perform()

                print(f"第 {i+1} 次点赞成功！")

        except Exception as e:
            print("点赞失败：", e)


if __name__ == "__main__":

    keyword = "123"
    start_page = 1
    end_page = 2
    max_items = 1
    insert_image = False

    scraper = Bilibili(keyword, start_page, end_page, insert_image,max_items=max_items)
    scraper.run()
