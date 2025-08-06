import sys
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image as PILImage

MAX_IMG_WIDTH = 80
MAX_IMG_HEIGHT = 80

class ExcelProcessor:
    def __init__(self, file_path, img_col_letter='L'):
        self.file_path = file_path
        self.img_col_letter = img_col_letter
        self.wb = None
        self.ws = None
        self.images = None
        self.data_list = []

    class Node:
        def __init__(self, data, image=None):
            self.data = data  # 存储行数据
            self.image = image  # 存储图片（ExcelImage对象）
            self.next = None

    def load_workbook(self):
        """加载 Excel 文件"""
        try:
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb.active
            self.images = self.ws._images  # 获取图片列表
            #print(f"成功加载文件: {self.file_path}")
            return True
        except Exception as e:
            #print(f"加载文件失败: {e}")
            return False

    def resize_image(self, img):
        """调整图片尺寸到最大宽高80*80，返回openpyxl的Image对象"""
        try:
            img_bytes = BytesIO(img._data())
            pil_img = PILImage.open(img_bytes)
            width, height = pil_img.size
            scale = min(MAX_IMG_WIDTH / width, MAX_IMG_HEIGHT / height, 1)
            new_size = (int(width * scale), int(height * scale))
            resized_img = pil_img.resize(new_size, PILImage.LANCZOS)

            output = BytesIO()
            resized_img.save(output, format='PNG')
            output.seek(0)

            return ExcelImage(output)
        except Exception as e:
            #print(f"图片缩放失败: {e}")
            return img

    def read_excel_to_linked_list(self):
        """从 Excel 读取数据并存入链表"""
        try:
            headers = [cell.value for cell in self.ws[1]]

            # 找到需要的列索引
            title_idx = headers.index("标题") if "标题" in headers else headers.index("Title")
            price_idx = headers.index("价格") if "价格" in headers else headers.index("Price")
            deal_idx = headers.index("成交量") if "成交量" in headers else headers.index("Deal")
            shop_url_idx = headers.index("店铺链接") if "店铺链接" in headers else headers.index("Title_URL")
            num_com_idx = headers.index("评论数量") if "评论数量" in headers else headers.index("Num_Com")
            shop_name_idx = headers.index("店铺名称") if "店铺名称" in headers else headers.index("Shop")
            is_post_free_idx = headers.index("包邮信息") if "包邮信息" in headers else headers.index("IsPostFree")
            tag_idx = headers.index("标签") if "标签" in headers else -1  # 没有则是空
            img_idx = headers.index("图片") if "图片" in headers else headers.index("Image")

            # 获取所有数据行
            data_rows = list(self.ws.iter_rows(min_row=2))

            # 将每行数据转换为链表节点
            for row in data_rows:
                new_row = [
                    row[title_idx].value,
                    row[price_idx].value,
                    row[deal_idx].value,
                    row[shop_url_idx].value,  # Shop_URL
                    row[shop_name_idx].value,
                    row[is_post_free_idx].value,
                    row[tag_idx].value if tag_idx != -1 else 'null',  # 标签，若无则是null
                    row[num_com_idx].value,
                ]

                # 处理图片（如果存在）
                img = self.images.pop(0) if len(self.images) > 0 else None
                if img:
                    new_img = self.resize_image(img)
                else:
                    new_img = None

                # 将行数据和图片存入链表节点
                new_node = self.Node(new_row, new_img)
                if not self.data_list:
                    self.data_list = new_node
                else:
                    # 找到链表的尾部插入
                    current = self.data_list
                    while current.next:
                        current = current.next
                    current.next = new_node

            #print("Excel 数据读取并存入链表完成")
            return True

        except Exception as e:
            #print(f"读取 Excel 数据失败: {e}")
            return False

    def remove_duplicates(self):
        """去除链表中重复的节点，比较标题和店铺名称"""
        seen = {}  # 用于存储已出现的 (标题, 店铺名称) 组合
        prev_node = None  # 用于指向前一个节点，方便删除重复节点
        current_node = self.data_list  # 从链表头开始

        while current_node:
            # 获取当前节点的标题和店铺名称
            title = current_node.data[0]  # 标题在列表中的位置
            shop_name = current_node.data[4]  # 店铺名称在列表中的位置
            key = (title, shop_name)

            if key in seen:
                # 如果 (标题, 店铺名称) 已经出现过，删除当前节点
                prev_node.next = current_node.next
            else:
                # 否则，将 (标题, 店铺名称) 添加到已见字典中
                seen[key] = True
                prev_node = current_node  # 更新前一个节点为当前节点

            # 移动到下一个节点
            current_node = current_node.next

        #print("去重完成")
        return self.data_list

    def filter_by_tag(self, mode, tag):
        """通过正则表达式过滤链表中的节点
        mode: bool (True/False) -- True 模式保留含 tag 的节点，False 模式保留不含 tag 的节点
        tag: str -- 用于匹配标题的标签文本
        """
        # 创建正则表达式模式，忽略大小写
        tag_pattern = re.compile(re.escape(tag), re.IGNORECASE)

        current_node = self.data_list
        prev_node = None

        while current_node:
            title = current_node.data[0]  # 获取标题数据（假设标题是列表中的第一个元素）

            # 正则匹配标题中的 tag
            match = tag_pattern.search(title)

            # mode=True：保留包含 tag 的节点，删除不包含 tag 的节点
            if mode:
                if match:
                    # 标题包含 tag，保留节点
                    prev_node = current_node  # 保持连接
                else:
                    # 标题不包含 tag，删除当前节点
                    if prev_node:
                        prev_node.next = current_node.next  # 使前一个节点指向下一个节点，删除当前节点
                    else:
                        # 如果是头节点，直接更新头节点
                        self.data_list = current_node.next

            # mode=False：保留不包含 tag 的节点，删除包含 tag 的节点
            else:
                if not match:
                    # 标题不包含 tag，保留节点
                    prev_node = current_node  # 保持连接
                else:
                    # 标题包含 tag，删除当前节点
                    if prev_node:
                        prev_node.next = current_node.next  # 使前一个节点指向下一个节点，删除当前节点
                    else:
                        # 如果是头节点，直接更新头节点
                        self.data_list = current_node.next

            # 移动到下一个节点
            current_node = current_node.next

        #print(f"过滤完成，{('保留包含', '保留不包含')[mode]}tag的节点")
        return self.data_list

    def write_linked_list_to_excel(self):
        """将链表数据写入新的 Excel 文件"""
        try:
            # 新表头
            new_headers = ["标题", "价格", "成交量", "店铺链接", "店铺名称", "包邮信息", "标签", "评论数量", "图片"]

            # 新表格初始化
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.append(new_headers)  # 写入新表头

            # 处理链表中的数据并写入 Excel
            current_node = self.data_list
            while current_node:
                row = current_node.data
                new_ws.append(row)

                # 如果节点有图片，插入图片
                if current_node.image:
                    anchor_cell = f"{get_column_letter(new_ws.max_column)}{new_ws.max_row}"
                    current_node.image.anchor = anchor_cell
                    new_ws.add_image(current_node.image)

                    # 调整行高和列宽
                    new_ws.row_dimensions[new_ws.max_row].height = 60
                    new_ws.column_dimensions[get_column_letter(new_ws.max_column)].width = 14

                current_node = current_node.next

            # 保存新文件
            output_path = os.path.splitext(self.file_path)[0] + "_转存.xlsx"
            new_wb.save(output_path)
            #print(f"保存完成: {output_path}")
            return True

        except Exception as e:
            #print(f"写入 Excel 文件失败: {e}")
            return False

    def Deduplication(self, file_path):
        """处理 Excel 文件"""
        if not os.path.isfile(file_path):
            #print("文件不存在:", file_path)
            return False

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in [".xlsx", ".xlsm"]:
            #print("仅支持 .xlsx 或 .xlsm 文件")
            return False

        self.file_path = file_path  # 将外部传入的 file_path 设置为实例的 file_path

        if not self.load_workbook():
            return False

        # 读取 Excel 数据并存入链表
        if not self.read_excel_to_linked_list():
            return False

        # 去除链表中的重复项
        self.remove_duplicates()

        # 将链表数据写入新的 Excel 文件
        success = self.write_linked_list_to_excel()
        return success

    def Screening(self, file_path, mode=True, tag="奉贤"):
        """处理 Excel 文件，根据 mode 和 tag 筛选链表中的节点"""
        #print ( file_path, mode, tag)
        if not os.path.isfile(file_path):
            #print("文件不存在:", file_path)
            return False

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in [".xlsx", ".xlsm"]:
            #print("仅支持 .xlsx 或 .xlsm 文件")
            return False

        self.file_path = file_path  # 将外部传入的 file_path 设置为实例的 file_path

        if not self.load_workbook():
            return False

        # 读取 Excel 数据并存入链表
        if not self.read_excel_to_linked_list():
            return False
        #print(mode, tag)
        # 过滤链表中的节点（根据用户输入的模式和 tag）
        self.filter_by_tag(mode, tag)

        # 将链表数据写入新的 Excel 文件
        success = self.write_linked_list_to_excel()
        return success

def main():
    # 修改为通过输入文件路径
    file_path = input("请输入 Excel 文件路径: ").strip()

    # 检查文件是否存在
    if not os.path.isfile(file_path):
        #print(f"文件 {file_path} 不存在，请重新输入有效的文件路径。")
        return

    processor = ExcelProcessor(file_path)

    success = processor.Screening()
    input("处理完成，按回车退出..." if success else "处理失败，按回车退出...")

if __name__ == "__main__":
    main()
