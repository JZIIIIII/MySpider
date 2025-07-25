# -*- coding: utf-8 -*-
from seleniumwire import webdriver
import undetected_chromedriver as uc
import random
import time
import os
import sys
import logging
import json
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.drawing.image import Image as ExcelImage
import hashlib
from path_utils import resource_path , static_log_path


class BaseScraper:
    def __init__(self, headless=True, proxy=None):
        self.logger = self._init_logger()  # 先初始化日志
        self.driver = self.init_driver(headless=headless, proxy=proxy)

    def _init_logger(self):
        logger = logging.getLogger("Spider")
        logger.setLevel(logging.DEBUG)  # 可调为 INFO 或 ERROR

        log_file_path = static_log_path("Mypider.log")

        file_handler = logging.FileHandler(log_file_path, encoding="utf-8")
        file_handler.setLevel(logging.DEBUG)

        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        if not logger.hasHandlers():
            logger.addHandler(file_handler)
            logger.addHandler(console_handler)

        return logger

    def human_sleep(self, min_time=0.5, max_time=1.5):
        time.sleep(random.uniform(min_time, max_time))        

    def save_page_html(self, filename="page_snapshot.html", wait_for_element=None, timeout=10):
        """ 保存当前页面的 HTML 内容到文件，确保页面完全加载 """
        try:
            if wait_for_element:
                WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, wait_for_element))
                )

            html_content = self.driver.page_source
            timestamp = time.strftime("%Y%m%d_%H%M%S")
        
            # 使用统一路径函数，拼接带时间戳的文件名
            file_path = static_log_path(f"{timestamp}_{filename}")

            with open(file_path, "w", encoding="utf-8") as file:
                file.write(html_content)

            self.logger.info(f"页面 HTML 已保存到: {file_path}")
        except Exception as e:
            self.logger.warning(f"保存页面 HTML 时出错: {e}")


    ##滑动窗口
    def scroll_step_down(self, base_step=300):
        """模拟人类向下小段滑动"""
        step = random.randint(base_step - 100, base_step + 100)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.6, 1.2))
        #self.wait_manual_verification()


    def smart_scroll_until_loaded(self, min_new_items=4, max_scroll_attempts=10):
        """边滚动边检测新商品是否加载，最多尝试 max_scroll_attempts 次"""
        last_count = 0
        attempts = 0

    def multi_scroll_up_down(self, up_num=1,down_num=8 , up_step=-200, down_step=800):
        # 向下滚动
        for _ in range(down_num):
            self.scroll_step_down(down_step)
    
        # 向上滚动
        for _ in range(up_num):
            self.scroll_step_down(up_step)
 

        # 模拟人工休眠，防止过快的滚动
        time.sleep(random.uniform(1.0, 1.5))

    def compute_hash(self, item: dict, platform_name: str ,features_url) -> str:
        """
        计算哈希值
        """
        title = item.get('title', '')
        price = item.get('price', '')

        base_str = f"{platform_name}|{title}|{price}|{features_url}"
        return hashlib.md5(base_str.encode('utf-8')).hexdigest()
    

    def load_hash_set(self,json_path: str) -> set:
        """
        从 JSON 文件中加载已处理过的哈希值集合
        """
        if not os.path.exists(json_path):
            return set()
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return set(data)

    def save_hash_set(self,hash_set: set, json_path: str):
        """
        将哈希集合保存为 JSON 文件
        """
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(list(hash_set), f, indent=2, ensure_ascii=False)

    def save_product_to_excel(self, row, title, price, deal, location, shop, post, item_url, shop_url, img_url, num_com, image_path=None):
        """
        保存商品信息到 Excel。

        :param row: 当前行数
        :param title: 商品标题
        :param price: 商品价格
        :param deal: 商品成交量
        :param location: 商品所在地
        :param shop: 店铺名称
        :param post: 包邮信息
        :param item_url: 商品链接
        :param shop_url: 店铺链接
        :param img_url: 图片链接
        :param num_com: 评论数量
        :param image_path: 图片路径（可选）
        """
        # 设置行高和列宽
        self.sheet.row_dimensions[row].height = 65
        self.sheet.column_dimensions['L'].width = 13

        # 填充数据到 Excel
        self.sheet.cell(row=row, column=1, value=row - 1)
        self.sheet.cell(row=row, column=2, value=title)
        self.sheet.cell(row=row, column=3, value=price)
        self.sheet.cell(row=row, column=4, value=deal)
        self.sheet.cell(row=row, column=5, value=location)
        self.sheet.cell(row=row, column=6, value=shop)
        self.sheet.cell(row=row, column=7, value=post)
        self.sheet.cell(row=row, column=8, value=item_url)
        self.sheet.cell(row=row, column=9, value=shop_url)
        self.sheet.cell(row=row, column=10, value=img_url)
        self.sheet.cell(row=row, column=11, value=num_com)

        # 插入图片
        if self.insert_image and image_path and os.path.exists(image_path):
            try:
                img = ExcelImage(image_path)
                img.width, img.height = 80, 80
                self.sheet.add_image(img, f'L{row}')
            except Exception as e:
                print(f"插入图片失败: {e}（图片路径: {image_path}）")

        print(f"第{row - 1}个商品信息已保存")    


    def init_driver(self, headless=None, proxy=None):
        # 使用 path_utils 的 resource_path
        chrome_path = resource_path("Chrome_Tool/GoogleChromePortable/GoogleChromePortable.exe")
        chromedriver_path = resource_path("Chrome_Tool/chromedriver-win64/chromedriver.exe")

        if not os.path.exists(chrome_path):
            self.logger.error("Chrome 浏览器路径不存在：%s", chrome_path)
            return None
        if not os.path.exists(chromedriver_path):
            self.logger.error("ChromeDriver 路径不存在：%s", chromedriver_path)
            return None

        options = uc.ChromeOptions()
        options.binary_location = chrome_path
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--start-maximized')

        if headless:
            options.add_argument('--headless=new')

        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
        ]

        options.add_argument(f"user-agent={random.choice(user_agents)}")

        resolutions = [(1920, 1080), (1366, 768), (1440, 900)]
        width, height = random.choice(resolutions)
        options.add_argument(f'--window-size={width},{height}')

        try:
            driver = uc.Chrome(
                options=options,
                driver_executable_path=chromedriver_path,
                use_subprocess=True
            )
            self.logger.info("ChromeDriver 启动成功")
        except Exception as e:
            self.logger.warning("Chrome 启动失败：%s", e)
            return None

        def interceptor(request):
            pass

        def response_interceptor(request, response):
            if 'react_psnl_verification_' in request.path:
                body = response.body.decode('utf-8', errors='ignore')
                modified = body.replace('navigator.webdriver', 'navigator.qwerasdfzxcv')
                response.body = modified.encode('utf-8')

        try:
            driver.request_interceptor = interceptor
            driver.response_interceptor = response_interceptor
            self.logger.info("请求/响应拦截器设置成功")
        except Exception as e:
            self.logger.warning("设置拦截器失败：%s", e)

        time.sleep(1)
        return driver
        

        
