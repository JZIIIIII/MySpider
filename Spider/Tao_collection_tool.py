# -*- coding: utf-8 -*-
# 代码说明：
'''
代码功能： 基于ChromeDriver爬取taobao（淘宝）平台商品列表数据
输入参数:  KEYWORLD --> 搜索商品“关键词”；
          pageStart --> 爬取起始页；
          pageEnd --> 爬取终止页；修改代码让图片获取的步骤在            deal = item.find('.realSales--XZJiepmt').text().replace("万", "0000").split("人")[0].split("+")[0]
            deal = int(deal) if deal.isdigit() else 0
            location = item.find('.procity--wlcT2xH9 span').text()
            shop = item.find('.shopNameText--DmtlsDKm').text()
            post = "包邮" if "包邮" in item.find('.subIconWrapper--Vl8zAdQn').text() else "/"
            item_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href')
            shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')
            img_url = item.find('.mainPicAdaptWrapper--V_ayd2hD img').attr('src')

            num_com = self.get_comment_count(item_url)此处实现
 
输出文件：爬取商品列表数据
        'Page'        ：页码
        'Num'         ：序号
        'title'       ：商品标题
        'Price'       ：商品价格
        'Deal'        ：商品销量
        'Location'    ：地理位置
        'Shop'        ：店铺
        'IsPostFree'  ：是否包邮
        'Title_URL'   ：商品详细页链接
        'Shop_URL'    ：商铺链接
        'Img_URL'     ：图片链接
        'Img'         ；图片
'''
from math import prod
import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from Spider.BaseScraper import BaseScraper
from selenium.common.exceptions import NoSuchElementException

from Spider.AntiScrapingException import CaptchaHandler
from path_utils import static_image_path ,static_hash_path , static_excel_path



class TaobaoScraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page,insert_image=True, max_items=100, pierce_span=[0,0]):
        super().__init__(headless=True, proxy=None)  # 调用父类初始化
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.insert_image = insert_image
        self.max_items = max_items
        self.pierce_span  = pierce_span
        self.all_warelists = []

        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self._setup_excel()
        self.hash_set = set()
        #self.anti_spider_triggered = False
        self.captcha_handler = CaptchaHandler(self.driver, self.logger)
        self.captcha_handler.Taosliderl()
        self.captcha_handler.TaoCheckRisk()




    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def save_to_excel(self, results, filename='results.xlsx'):
        # 创建新的 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = '商品信息'

        # 初始化表头
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']

        # 写入表头
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)

        # 设置图片插入列宽
        img_col_letter = 'L'  # 图片插入列是第L列
        ws.column_dimensions[img_col_letter].width = 15

        # 填充商品数据
        for i, item in enumerate(results, start=2):  # 从第二行开始插入数据
            # 填充商品信息
            ws.cell(row=i, column=1, value=i - 1)  # 商品编号（Num）
            ws.cell(row=i, column=2, value=item['title'])  # 商品标题（Title）
            ws.cell(row=i, column=3, value=item['price'])  # 商品价格（Price）
            ws.cell(row=i, column=4, value=item['deal_num'])  # 成交量（Deal）
            ws.cell(row=i, column=5, value=item['location'])  # 商品所在地（Location）
            ws.cell(row=i, column=6, value=item['shop_name'])  # 店铺名称（Shop）
            ws.cell(row=i, column=7, value=item['freeFreight'])  # 是否包邮（IsPostFree）
            ws.cell(row=i, column=8, value=item['item_url'])  # 商品链接（Title_URL）
            ws.cell(row=i, column=9, value=item['shop_url'])  # 店铺链接（Shop_URL）
            ws.cell(row=i, column=10, value=item['img_url'])  # 图片链接（Img_URL）
            ws.cell(row=i, column=11, value=item['comment'])  # 评论数量（Num_Com）

            # 插入商品图片（Image）
            if self.insert_image:
                img_path = item.get('img_path')  # 商品图片路径
                if img_path and os.path.exists(img_path):
                    try:
                        # 使用 Pillow 打开图片并重新编码为 JPEG（修复部分 JPEG 插入失败问题）
                        with Image.open(img_path) as img:
                            rgb_img = img.convert("RGB")
                            rgb_img.save(img_path, format='JPEG')  # 覆盖原图

                        # 插入图片到 Excel
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 80
                        excel_img.height = 80
                        ws.row_dimensions[i].height = 60  # 设置行高
                        excel_img.anchor = f'{img_col_letter}{i}'  # 图片插入到指定单元格
                        ws.add_image(excel_img)
                    except Exception as e:
                        self.logger.error(f"[!] 图片插入失败: {img_path} | 错误: {e}")
                else:
                    self.logger.warning(f"[!] 图片路径无效或文件不存在: {img_path}")

        # 保存 Excel 文件
        wb.save(filename)
        self.logger.info(f"[√] 数据已保存到 Excel：{filename}")


    def save_cookies(self, path="taobao_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="taobao_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):
        #print("正在尝试登录淘宝...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)

            input("如出现滑块，请手动完成验证后按 Enter 继续...")

            self.save_cookies()
        except Exception as e:
            self.logger.warning(f"登录失败: {e}")
            self.save_page_html("error_page.html")

    def search(self):
        try:
            self.logger.info(f"尝试用s方法定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'q')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="J_SearchForm"]/div/div[1]/button')))
        except Exception as e1:
            self.logger.warning(f"s方法失败: {e1}\n尝试用w方法定位...")
            try:
                search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'search_input')))
                search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'search-btn')))
            except Exception as e2:
                self.logger.warning(f"w方法也失败: {e2}")
                try:
                    # 新版页面方法
                    search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'q')))
                    search_btn = self.wait.until(EC.element_to_be_clickable((By.ID, 'button')))
                except Exception as e3:
                    self.logger.error(f"search？搜索框定位也失败: {e3}")
                    self.save_page_html("error_page.html")
                    raise Exception("搜索框和按钮定位失败，搜索终止")
                s
                raise Exception("搜索框和按钮定位失败，搜索终止")

                self.save_page_html("error_page.html")

        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.human_sleep(2, 3)
        except Exception as e:
            self.logger.error(f"搜索操作失败:{e}")
            self.save_page_html("error_page.html")

    def go_to_page(self, page_number):
        try:
            # === 风控检查与等待控制 ===
            self.RiskPause(self.captcha_handler.Taosliderl())
            self.RiskPause(self.captcha_handler.TaoCheckRisk())
            # 等待跳页输入框出现
            page_input = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'span.next-input input')))
            page_input.clear()
            page_input.send_keys(str(page_number))
            self.human_sleep()

            # 找到并点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button.next-pagination-jump-go')))
            confirm_btn.click()

            # 等待当前页码变为目标页
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.next-pagination-display em'),
                str(page_number)))

            self.human_sleep(3, 4)

        except Exception as e:
            self.logger.error(f"翻页失败: {e}")
            self.save_page_html("error_page.html")
            
    def get_comment_count(self, item_id): 
        """
        打开商品详情页并提取评论数。
        支持风控处理与暂停/终止控制。
        """
        count = 0
        deal_text = ''
        main_window = self.driver.current_window_handle

        try:
            # 精确定位商品 <a> 标签，避免重复点击第一个 item
            link_xpath = f'//a[@data-spm-act-id="{item_id}"]'
            link_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, link_xpath)))

            # 滚动到可见并点击
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link_element)
            self.human_sleep(0.5, 1)
            link_element.click()
            self.human_sleep(1, 2)

            # 等待新标签页打开
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return 0, 0

            # 切换到新标签页
            new_tab = [h for h in self.driver.window_handles if h != main_window][-1]
            self.driver.switch_to.window(new_tab)
            self.human_sleep(2, 3)

            # === 风控检查与等待控制 ===
            self.RiskPause(self.captcha_handler.Taosliderl())
            self.RiskPause(self.captcha_handler.TaoCheckRisk())

            self.wait_if_paused()
            if self.should_stop():
                self.logger.info("用户主动终止爬虫，跳过当前评论数解析")
                self.stop()
                return 0, 0

            # ===== 以下保持你原有的销量和评论获取逻辑 =====
            try:
                # 等待销量模块加载
                self.wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]')
                ))

                deal_text = self.driver.find_element(
                    By.XPATH,
                    '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]'
                ).text.strip()

                if not deal_text:
                    deal_text = self.driver.execute_script(
                        "return arguments[0].textContent;", 
                        self.driver.find_element(
                            By.XPATH,
                            '//div[contains(@class,"salesDesc--Z35wP98o") or contains(@class,"salesDesc--mOo6I5Bc")]'
                        )
                    ).strip()

                self.logger.info(f"获取销量文本: {deal_text}")
                deal_text = deal_text.lstrip('·').strip()

            except Exception as e:
                self.logger.warning(f"获取销量失败: {e}")
                deal_text = "商品详情获取销量失败 可能为0"

            try:
                # 等待至少一个「用户评价」标题出现（注意：现在是 div 不是 span）
                self.wait.until(lambda d: d.find_elements(
                    By.CSS_SELECTOR, "div.tabDetailItemTitle--bJtPXTNu"
                ))

                comment_nodes = self.driver.find_elements(
                    By.CSS_SELECTOR, "div.tabDetailItemTitle--bJtPXTNu"
                )

                count = 0
                for node in comment_nodes:
                    text = node.text.strip()
                    # text 示例："用户评价 · 100+"
                    if text.startswith("用户评价"):
                        try:
                            # 用 · 分割，取最后一段
                            num_part = text.split('·')[-1].strip()
                            num_part = num_part.replace('+', '').replace(',', '')

                            if '万' in num_part:
                                count = int(float(num_part.replace('万', '')) * 10000)
                            else:
                                count = int(num_part)
                        except Exception as e:
                            self.logger.warning(f"评论数解析失败: {text} | {e}")
                            count = 0
                        break  # 找到就退出

            except TimeoutException:
                self.logger.warning("评论模块加载超时，评论数设为 0")
                count = 0

        except Exception as e:
            self.logger.warning(f"获取详情失败：{e}")
            self.save_page_html("error_page.html")

        finally:
            # 关闭新标签页并返回主窗口
            self.human_sleep(1, 2)
            try:
                if self.driver.current_window_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
            except Exception as e:
                self.logger.error(f"关闭标签页或切换窗口失败：{e}")
                self.save_page_html("error_page.html")
    
        return deal_text, count

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.taobao.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(f"图片请求失败: {url}")
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(f"图片解码失败（格式可能不受支持，如webp）: {url}")
                return None
            except Exception as e:
                self.logger.error(f"图片处理异常: {e} | URL: {url}")
                return None

            # 使用硬编码路径创建文件夹
            folder = "images/TaoBao"  # 直接使用硬编码的路径
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 构建文件名和文件路径
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(f"图片下载异常: {e} | URL: {url}")
            return None

    def download_all_images(self):
        """
        批量下载 self.all_warelists 中的图片并保存到本地。
        并在每个商品字典中添加 'image_path' 字段。
        """
        for index, product_info in enumerate(self.all_warelists):
            img_url = product_info.get('img_url', 'N/A')

            if img_url != 'N/A':  # 如果图片链接有效
                image_path = self.download_image(img_url, index + 1)  # 下载并保存图片
                if image_path:
                    product_info['img_path'] = image_path  # 更新商品字典，保存图片的本地路径
                else:
                    product_info['img_path'] = None  # 下载失败则设为 None
            else:
                product_info['img_path'] = None  # 如果没有图片链接，也设为 None

        self.logger.info(f"已成功下载并保存 {len(self.all_warelists)} 张商品图片")


    def clean_image_folder(self):
        folder = "images/TaoBao"  # 直接使用硬编码的路径
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # 删除文件或快捷方式
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # 删除子文件夹


    def JSlistener(self, driver):

        script = """
        (function() {

            // =========================
            // 防重复注入
            // =========================
            if (window.__TB_JSONP_FIXED__) return;
            window.__TB_JSONP_FIXED__ = true;

            // =========================
            // 内存桥（替代 localStorage）
            // =========================
            window.__TB_BRIDGE__ = window.__TB_BRIDGE__ || [];

            function push(data, url) {
                window.__TB_BRIDGE__.push({
                    t: Date.now(),
                    url: url,
                    data: data
                });
            }

            // =========================
            // 核心：hook callback
            // =========================
            function hook(name, url) {

                if (!name) return;

                const old = window[name];

                if (old && old.__hooked__) return;

                window[name] = function(data) {

                    push(data, url);

                    return typeof old === 'function'
                        ? old.apply(this, arguments)
                        : undefined;
                };

                window[name].__hooked__ = true;
            }

            // =========================
            // 关键：拦截 script 插入
            // =========================
            const origAppend = Element.prototype.appendChild;

            Element.prototype.appendChild = function(el) {

                try {

                    if (el && el.tagName === 'SCRIPT' && el.src) {

                        const m = el.src.match(/[?&]callback=([^&]+)/);

                        if (m) {
                            const cb = decodeURIComponent(m[1]);

                            // ⚠️ 关键：每次都 hook（不依赖旧状态）
                            hook(cb, el.src);
                        }
                    }

                } catch (e) {}

                return origAppend.call(this, el);
            };

        })();
        """

        driver.execute_script(script)



    def parse_page(self, page_number, platform='TaoBao', hash_json=None):
        count_number = len(self.all_warelists)
        #暂停检测
        self.wait_if_paused()
        #终止检测
        if self.should_stop():
            self.logger.info("检测到终止命令")
            return 'finish'  # 返回 'finsih' 表示抓取终止
        try:
            # 1. 只保证 hook 存在（不重复注入也没关系）
            self.JSlistener(self.driver)

            # 2. 只清内存桥（关键）
            self.driver.execute_script("""
                window.__TB_BRIDGE__ = [];
            """)

            self.human_sleep(1, 2)

            # 3. 跳页
            if page_number == self.page_start:
                self.go_to_page(page_number+1)
                self.human_sleep(1, 2)
                self.driver.execute_script("""
                    window.__TB_BRIDGE__ = [];
                """)
                self.human_sleep(1, 2)
                self.go_to_page(page_number)
            else:
                self.go_to_page(page_number)

            self.human_sleep(1, 2)

            self.multi_scroll_up_down(0, 8, 0, 600)

            self.human_sleep(0.5, 1)

            # 4. 取数据
            saved_data = self.driver.execute_script("""
                const d = window.__TB_BRIDGE__ || [];
                window.__TB_BRIDGE__ = [];
                return d;
            """)

            #print(data)
            #print(f"第 {page_number} 页数据条数: {len(data)}")
            #with open("tb_search_data.json", "w", encoding="utf-8") as f:
            #  json.dump(saved_data, f, ensure_ascii=False, indent=4)

                        # 正向遍历 saved_data 列表
            for item in saved_data:
                if isinstance(item, dict) and isinstance(item.get("data"), dict):
                    data = item["data"]
                    if ".recommend" not in data.get("api", ""):
                        continue
                    target_data = data.get("data")
                    if isinstance(target_data, dict) and "itemsArray" in target_data:
                        itemsArray = target_data.get("itemsArray", [])

                        for product in itemsArray:
                            if not (product.get("title") and product.get("price")):
                                continue
                            product_info = {
                                'title': product.get('title', 'N/A'),
                                'price': product.get('price', 'N/A'),
                                'deal_num': product.get('realSales', 'N/A'),
                                'location': product.get('procity', 'N/A'),
                                'shop_name': product.get('nick', 'N/A'),
                                'item_url': product.get('auctionURL', 'N/A'),
                                'shop_url': (product.get('shopInfo') or {}).get('url', 'N/A'),
                                'comment': "",
                                'img_url': product.get('pic_path', 'N/A'),
                                'freeFreight': "包邮" if any(
                                    i.get("text") == "包邮"
                                    for i in (product.get("icons") or [])
                                ) else ""
                            }
                            # title 清洗
                            title = product_info['title']
                            title = self.remove_html_tags(title)
                            title = self.decode_html_entities(title)
                            product_info['title'] = title
                            # URL 补全
                            item_url = product_info['item_url']
                            if isinstance(item_url, str) and item_url.startswith('//'):
                                item_url = 'https:' + item_url
                            product_info['item_url'] = item_url

                            shop_url = product_info['shop_url']
                            if isinstance(shop_url, str) and shop_url.startswith('//'):
                                shop_url = 'https:' + shop_url
                            product_info['shop_url'] = shop_url

                            img_url = product_info['img_url']
                            if isinstance(img_url, str) and img_url.startswith('//'):
                                img_url = 'https:' + img_url
                            product_info['img_url'] = img_url
                            # ==== 价格过滤 ==== 
                            min_p, max_p = self.pierce_span

                            # 防止用户输入反区间 [100, 50]
                            if min_p > max_p:
                                min_p, max_p = max_p, min_p

                            # 只有当区间不是 [0,0] 时才进行过滤
                            if not (min_p == 0 and max_p == 0):
                                try:
                                    # 获取商品价格（从 `product_info['price']` 获取）
                                    price_value = float(product_info['price']) if product_info['price'] != 'N/A' else 0
                                except:
                                    self.logger.warning(f"无法解析价格 {product_info['price']}")
                                    continue  # 无法解析价格 → 跳过

                                # 不在闭区间 → 跳过该商品
                                if not (min_p <= price_value <= max_p):
                                    self.logger.info(f"[价格过滤] {price_value} 不在区间 [{min_p}, {max_p}] 内，跳过")
                                    continue

                            # ==== 哈希去重 ====
                            # 使用 compute_hash 方法生成商品的哈希值
                            item_dict = {'title': product_info['title'], 'price': product_info['price']}
                            features_url = product_info['item_url']
                            product_hash = self.compute_hash(item_dict, 'JD', features_url)

                            # 如果商品哈希值已经存在，跳过
                            if product_hash in self.hash_set:
                                self.logger.info(f"[跳过] 已存在商品: {product_info['title']}")
                                continue

                            # 添加哈希值到集合
                            self.hash_set.add(product_hash)

                            # 将商品信息添加到当前页商品列表
                            self.all_warelists.append(product_info)
                            count_number += 1
                            self.count += 1
                            self.logger.info(f"提取到商品数量: {count_number}")

                            # 如果已经抓取到最大数量的商品，停止处理
                            if count_number >= self.max_items:
                                self.logger.info(f"已抓取 {self.max_items} 条商品数据，停止抓取此页数据...")
                                return 'max_reached'  # 返回 'max_reached' 表示达到最大数量


            return 'continue'  # 返回 'continue' 表示数据抓取成功并继续抓取下一页


        except Exception as e:
            # 记录详细的错误信息
            self.logger.error(f"提取商品信息失败: 第{page_number}页, 错误: {str(e)}")
            return 'error'  # 返回 'error' 表示抓取失败，外层循环可以决定是否继续


    def turn_page(self, page_number):
        try:
            # 找到“下一页”按钮（根据 class）
            next_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button.next-next:not([disabled])')))
        
            next_btn.click()

            # 等待当前页码显示变为指定页数
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.next-pagination-display em'),
                str(page_number)))

            self.human_sleep(2, 3)

        except Exception as e:
            self.logger.error(f"下一页点击失败: {e}")


    def run(self):
        try:

            self.driver.get("https://s.taobao.com/")


            self.search()

            # 等待前端登录完成
            self.wait_for_login()
            #input("如出现滑块，请手动完成验证后按 Enter 继续...")

            # 初始化抓取的商品信息列表
            self.all_warelists = []

            # 循环遍历指定的页面范围
            for page_number in range(self.page_start, self.page_end + 1):

                #暂停检测
                self.wait_if_paused()
                #终止检测
                if self.should_stop():
                    self.logger.info("检测到终止命令")
                    break

                self.logger.info(f"开始抓取第 {page_number} 页的数据...")

                # 调用 parse_page 进行数据抓取
                result = self.parse_page(page_number)
                # 根据返回的结果决定是否继续抓取
                if result == 'max_reached':
                    self.logger.info(f"已抓取 {self.max_items} 条商品数据，停止抓取...")
                    break  # 达到最大数量，停止抓取
                elif result == 'no_data':
                    self.logger.info(f"第 {page_number} 页没有数据，跳过...")
                    break  # 没有数据终止
                elif result == 'error':
                    self.logger.error(f"第 {page_number} 页抓取失败，跳过...")
                elif result == 'finish':
                    break  # 发生错误终止
                self.human_sleep(1, 2)

            # 打印或处理抓取的商品数据
            self.logger.info(f"共抓取到 {len(self.all_warelists)} 条商品数据")

            data = self.all_warelists[:self.max_items]  # 只保留前 max_items 条数据
            if data:
                # 使用硬编码路径
                filename = f"{self.keyword}_{time.strftime('%Y%m%d_%H%M')}.xlsx"
                folder = "excel"  # 硬编码路径
                if not os.path.exists(folder):
                    os.makedirs(folder)  # 如果目录不存在，则创建

                # 生成完整的文件路径
                filepath = os.path.join(folder, filename)
                self.download_all_images()  # 批量下载并更新图片路径
                self.save_to_excel(data, filepath)  # 保存数据到 Excel
            self.clean_image_folder()  # 清理图片缓存


        except Exception as e:
            # 记录详细的错误信息
            self.logger.error(f"发生错误：{e}")

        self.driver.close()
        self.driver.quit()


if __name__ == "__main__":
    keyword = input("输入搜索关键词：")
    start_page = int(input("起始页码："))
    end_page = int(input("终止页码："))
    max_items = int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"
    pierce_span=[0,100]

    scraper = TaobaoScraper(keyword, start_page, end_page, insert_image,max_items=max_items ,pierce_span=pierce_span)
    scraper.run()
