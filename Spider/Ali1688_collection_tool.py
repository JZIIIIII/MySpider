# -*- coding: utf-8 -*-
import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from Spider.BaseScraper import BaseScraper
from Spider.AntiScrapingException import CaptchaHandler
from path_utils import static_image_path ,static_hash_path , static_excel_path



class Ali_1688Scraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, insert_image = True, max_items=100):
        super().__init__(headless=True, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self.insert_image = insert_image
        self._setup_excel()
        self.hash_set = set()
        #self.lock = threading.Lock()  # 用于同步的锁
        self.anti_spider_triggered = False
        # 创建 CaptchaHandler 实例并调用 wait_for_slider_manual 方法
        self.captcha_handler = CaptchaHandler(self.driver, self.logger)
        self.captcha_handler.AliCaptcha()  # 调用风控


    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    '''
    def save_cookies(self, path="Ali1688_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="Ali1688_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):
        print("正在尝试登录1688...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)
            input("如出现滑块，请手动完成验证后按 Enter 继续...")
            self.save_cookies()
        except Exception as e:
            self.logger.warning(f"登录失败:{e}")
    '''
    def search(self):
        try:
            self.logger.info("尝试用定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'alisearch-input')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'input-button')))
            self.RiskPause(self.captcha_handler.AliCaptcha())
        except Exception as e:
            self.logger.info("搜索框定位失败")
        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.RiskPause(self.captcha_handler.AliCaptcha())
            self.human_sleep(2, 3)
        except Exception as e:
            self.logger.warning(f"搜索操作失败:{e}")
            self.save_page_html("error_page.html")

    def go_to_page(self, page_number):
        try:
            # 定位输入框，清空并输入页码
            page_input = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'input.input-page')))
            page_input.clear()
            page_input.send_keys(str(page_number))
            self.human_sleep(0.5, 1.2)

            # 定位并点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'div.paging-to-page-button')))
            confirm_btn.click()

            self.RiskPause(self.captcha_handler.AliCaptcha())
            self.human_sleep(2, 3)

            return None  # 表示跳转成功

        except Exception as e:
            self.logger.warning(f"[翻页失败] 第 {page_number} 页: {e}")
            self.save_page_html("error_page.html")

            # ==== 提取最大页码 ====
            try:
                page_info_elem = self.driver.find_element(By.CSS_SELECTOR, 'span.fui-paging-num')
                max_page = int(page_info_elem.text.strip())
                self.logger.warning(f"[检测最大页数] 当前最大页数为: {max_page}")
                return max_page
            except Exception as ex:
                self.logger.warning(f"[获取最大页码失败]: {ex}")
                return 1  # 安全兜底：默认只有 1 页

    def get_more(self, url):
        comment_count = 0
        location = ''
        combined_text = ''
        main_window = self.driver.current_window_handle

        try:
            # 提取商品ID，防止href匹配失败
            item_id_match = re.search(r'offerId=(\d+)', url)
            if not item_id_match:
                self.logger.warning("无法提取 offerId，跳过")
                return location, comment_count
            item_id = item_id_match.group(1)

            # 🔹新版页面的详情链接 (detail.m.1688.com)
            title_xpath = f'//a[contains(@href, "offerId={item_id}") and contains(@href, "detail.m.1688.com")]'
            title_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, title_xpath)))

            href = title_element.get_attribute('href')
            # self.logger.info(f"即将点击链接：{href}")
            #print(f"即将点击链接：{href}")

            # 获取尺寸并设置更合适的偏移
            size = title_element.size
            width = size['width']
            height = size['height']
            offset_x = int(width * 0.1)
            offset_y = int(height * 0.3)

            # 执行偏移点击
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL) \
                   .move_to_element_with_offset(title_element, offset_x, offset_y) \
                   .click() \
                   .key_up(Keys.CONTROL) \
                   .perform()

            self.RiskPause(self.captcha_handler.AliCaptcha())
            # 等待新标签页打开，最多等待5秒
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return location, comment_count

            # 切换到新标签页
            handles = self.driver.window_handles
            new_tabs = [h for h in handles if h != main_window]
            if not new_tabs:
                self.logger.warning("找不到新标签页句柄")
                self.save_page_html("error_page.html")
                return location, comment_count

            new_tab = new_tabs[-1]
            self.driver.switch_to.window(new_tab)
            self.RiskPause(self.captcha_handler.AliCaptcha())
            self.wait_if_paused()
            if self.should_stop():
                self.logger.info("用户主动终止爬虫，跳过当前评论数解析")
                self.stop()
                return location,comment_count

            self.human_sleep(2, 3)

            try:
                trade_info = self.driver.find_element(By.CSS_SELECTOR, "div.trade-info.v-flex")
                texts = self.driver.execute_script("""
                    let trade = arguments[0];
                    let nodes = trade.childNodes;
                    let texts = [];
                    for(let i=0; i<nodes.length; i++) {
                        let node = nodes[i];
                        if(node.nodeType === 3) { // 文本节点
                            let t = node.textContent.trim();
                            if(t) texts.push(t);
                        } else if(node.nodeType === 1) { // 元素节点
                            texts.push(node.innerText.trim());
                        }
                    }
                    return texts;
                """, trade_info)
                # 拼接从第二个元素开始的所有文本（过滤空）
                combined_text = ''.join(t for t in texts[1:] if t)
            except Exception as e:
                self.logger.warning(f"获取 trade-info 文本失败: {e}")
                combined_text = ""  # 失败则为空字符串或你需要的默认值

            # 后续用 combined_text 即可
            self.logger.warning(f"拼接结果:{combined_text}")

            # 获取评论数
            try:
                element = self.driver.find_element(By.CSS_SELECTOR, 'span.brackets')
                comment_count = int(element.get_attribute("data-value"))
            except Exception as e:
                self.logger.warning(f"获取评论数失败，错误信息：{e}")
                comment_count = 0

            # 获取发货地
            try:
                location_element = self.driver.find_element(By.CSS_SELECTOR, 'span.location')
                location = location_element.text.strip()
            except Exception as e:
                self.logger.warning(f"获取发货地失败，错误信息：{e}")
                location = ""

        except Exception as e:
            self.logger.warning(f"标签页打开失败，错误信息：{e}")

        finally:
            # 关闭新标签页，切换回主窗口
            self.human_sleep(1, 2)
            try:
                current_handle = self.driver.current_window_handle
                if current_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    self.RiskPause(self.captcha_handler.AliCaptcha())

            except Exception as e:
                self.logger.warning(f"关闭标签页或切换主窗口失败，错误信息：{e}")
                self.save_page_html("error_page.html")

        return combined_text,location, comment_count

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.1688.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(f"图片请求失败: {url}")
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(f"图片解码失败（格式可能不受支持，如webp）: {url}")
                return None
            except Exception as e:
                self.logger.error(f"图片处理异常: {e} | URL: {url}")
                return None

            # 使用硬编码路径
            folder = "images/1688"  # 硬编码文件夹路径
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 生成文件名
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)  # 拼接路径

            # 保存为 jpg 文件
            img.save(file_path, format="JPEG")

            return file_path

        except Exception as e:
            self.logger.error(f"图片下载异常: {e} | URL: {url}")
            return None

    def clean_image_folder(self):
        folder = "images/1688"  # 使用硬编码路径获取文件夹
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)  # 删除文件或快捷方式
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)  # 删除子文件夹
                except Exception as e:
                    self.logger.warning(f"[!] 删除缓存文件失败: {file_path} | 错误: {e}")
        else:
            self.logger.error(f"[!] 文件夹不存在或不是目录：{folder}")


    def parse_page(self, page_number, platform='Ali1688',hash_json=None):
        # 获取当前页面所有商品元素（Selenium WebElement）
        items_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a.search-offer-item')
        # 用 pyquery 解析页面HTML，拿到所有商品节点
        html = self.driver.page_source
        doc = pq(html)
        items_data = list(doc('a.search-offer-item').items())
        # 载入历史哈希
        if hash_json:
            hash_json = static_hash_path(hash_json)
            hash_set = self.load_hash_set(hash_json)
            self.hash_set.update(hash_set)  # 合并新的哈希池
        else:
            hash_set = self.hash_set  # 使用共享的哈希池
        slide_counter = 0  # 下滑计数器

        if len(items_elements) != len(items_data):
            self.logger.warning(f"警告：Selenium 找到商品数量 {len(items_elements)} 与 pyquery 找到数量 {len(items_data)} 不一致，可能页面未完全加载！")
            # 这里可以等待或重试，这里先简单返回False停止


        for idx, (element, data) in enumerate(zip(items_elements, items_data)):
            if self.count - 2 >= self.max_items:
                self.logger.info(f"已达到最大抓取数量：{self.max_items}，停止抓取。")
                if hash_json:
                    self.save_hash_set(self.hash_set, hash_json)  # 使用类成员变量 self.hash_set  
                return False

            # 如果暂停，则等待恢复
            self.wait_if_paused()

            # 检查是否被强制终止
            if self.should_stop():
                self.logger.info("检测到提前终止命令，保存已抓取内容并退出 parse_page。")
                if hash_json:
                    self.save_hash_set(self.hash_set, hash_json)  # 使用类成员变量 self.hash_set  
                return False

            try: 


                '''
                if self.anti_spider_triggered == True:
                    print(f"触发强制风控账号冻结 应前往1688解封")
                    self.save_hash_set(hash_set, hash_json)   
                    return False
                '''

                # 滚动到当前商品，使其位于中央
                self.driver.execute_script("""
                    var elem = arguments[0];
                    var rect = elem.getBoundingClientRect();
                    var windowHeight = window.innerHeight;
                    var elementTop = rect.top;
                    var elementBottom = rect.bottom;
                    var scrollAmount = (elementTop + elementBottom) / 2 - windowHeight / 2;
                    window.scrollBy(0, scrollAmount);
                """, element)

                # 解析图片URL
                img_tag = data.find('.offer-img-inner img') or data.find('img')
                if not img_tag:
                    img_tag = data.find('img')

                img_url = ""
                if img_tag:
                    img_url = (
                        img_tag.attr('src') 
                        or img_tag.attr('data-src') 
                        or img_tag.attr('data-lazyload') 
                        or ""
                    ).strip()

                if img_url.startswith("//"):
                    img_url = "https:" + img_url
                elif img_url.startswith("/"):
                    img_url = "https://cbu01.alicdn.com/" + img_url



                # 解析其他字段
                title = data.find('.offer-title-row .title-text div').text()

                price_container = data.find('.price-item')
                if price_container:
                    # 用集合记录是否遇到 '¥'
                    seen_yuan = set()  # 哈希表来记录是否遇到 '¥'
                    price_parts = []
    
                    for child in price_container.children():
                        part = pq(child).text().strip()
                        if part:  # 排除空白内容
                            # 如果遇到 '¥' 符号并且它已经出现过，停止继续添加
                            if '¥' in part:
                                if '¥' in seen_yuan:
                                    break  # 如果 '¥' 已经出现过一次，停止处理
                                seen_yuan.add('¥')  # 记录第一次遇到 '¥'
                            price_parts.append(part)

                    # 合并前两个价格部分
                    price_text = ''.join(price_parts[:2]).replace('¥', '').strip()

                    # 处理价格转换并控制小数点后两位
                    try:
                        price = round(float(price_text), 2)
                    except ValueError:
                        price = 0.0
                else:
                    price = 0.0
                '''
                try:
                    desc_text_elem = data.find('.col-desc_after .offer-desc-item .desc-text')
                    deal_text = desc_text_elem.text().strip() if desc_text_elem else "0"

                    if deal_text:
                        match = re.search(r'([\d\.]+)(万)?', deal_text)
                        if match:
                            number_part = match.group(1)
                            wan_unit = match.group(2)
                            deal = int(float(number_part) * (10000 if wan_unit else 1))
                        else:
                            deal = 0
                    else:
                        deal = 0
                except Exception:
                    deal = 0
                '''

                desc_text_elems = data.find('.col-desc').find('.desc-text')
                texts = [elem.text().strip() for elem in desc_text_elems.items()]
                post = "包邮" if any("包邮" in t for t in texts) else "/"

                try:
                    shop_elem = data.find('.col-left a.offer-desc-item')
                    shop = shop_elem.find('.desc-text').text().strip() if shop_elem else ''
                    shop_url = shop_elem.attr('href').strip() if shop_elem else ''
                    if shop_url.startswith("//"):
                        shop_url = "https:" + shop_url
                except Exception:
                    shop = ''
                    shop_url = ''

                item_url = data.attr('href')
                if item_url.startswith("//"):
                    item_url = "https:" + item_url
                elif item_url.startswith("/"):
                    item_url = "https://detail.1688.com" + item_url

                # ==== 新增哈希去重 ====
                item_dict = {'title': title, 'price': price}
                features_url = shop_url or ''  # 用图片链接代替特征链接

                item_hash = self.compute_hash(item_dict, platform, features_url)
                if item_hash in hash_set:
                    # print(f"[跳过] 已存在的商品: {title}")
                    continue
                self.hash_set.add(item_hash)  # 更新哈希池
                # ======================
                # 判断图片链接是否有效并下载
                if not img_url or not isinstance(img_url, str) or not img_url.startswith(('http', '//', '/')):
                    self.logger.warning(f"[警告] 第 {self.count - 1} 个商品图片 URL 无效，跳过插图")
                    self.logger.warning("商品 HTML 片段：")
                    self.logger.info(data.outer_html())
                    image_path = None
                else:
                    image_path = self.download_image(img_url, self.count - 1) if self.insert_image else None

                # 获取额外信息（地区、评论数）
                deal , location, num_com = self.get_more(item_url)

                # 写入 Excel
                row = self.count
                self.sheet.row_dimensions[row].height = 65
                self.sheet.column_dimensions['L'].width = 13

                self.sheet.cell(row=row, column=1, value=row - 1)
                self.sheet.cell(row=row, column=2, value=title)
                self.sheet.cell(row=row, column=3, value=price)
                self.sheet.cell(row=row, column=4, value=deal)
                self.sheet.cell(row=row, column=5, value=location)
                self.sheet.cell(row=row, column=6, value=shop)
                self.sheet.cell(row=row, column=7, value=post)
                self.sheet.cell(row=row, column=8, value=item_url)
                self.sheet.cell(row=row, column=9, value=shop_url)
                self.sheet.cell(row=row, column=10, value=img_url)
                self.sheet.cell(row=row, column=11, value=num_com)

                # 插入图片
                if self.insert_image and image_path and os.path.exists(image_path):
                    try:
                        img = ExcelImage(image_path)
                        img.width, img.height = 80, 80
                        self.sheet.add_image(img, f'L{row}')
                    except Exception as e:
                        self.logger.error(f"插入图片失败: {e}（图片路径: {image_path}）")

                self.logger.info(f"第{row - 1}个商品信息已保存")
                self.count += 1
                slide_counter += 1
                '''
                if slide_counter % 6 == 0:
                    self.scroll_step_down()
                self.human_sleep(1, 2)
                '''
            except Exception as e:
                self.logger.warning(f"[!] 解析商品异常: {e}")
                continue      

        # 保存哈希集，方便下次增量爬取
        if hash_json:
            self.save_hash_set(self.hash_set, hash_json)  # 使用类成员变量 self.hash_set
        return True

    def turn_page(self, page_number: int):
        try:
            # 1. 定位输入框并输入目标页码
            input_box = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input.input-page")))
            input_box.clear()
            input_box.send_keys(str(page_number))
            self.human_sleep(0.3, 1.0)

            # 2. 点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "div.paging-to-page-button")))
            confirm_btn.click()

            # 3. 滑块验证
            self.RiskPause(self.captcha_handler.AliCaptcha())

            # 4. 等待翻页完成，适配不同结构的当前页码元素
            def page_loaded(driver):
                try:
                    current_elem = driver.find_element(By.CSS_SELECTOR, "div.fui-current.fui-page-item")
                    return current_elem.text.strip() == str(page_number)
                except Exception:
                    return False

            WebDriverWait(self.driver, 10).until(page_loaded)

            self.logger.info(f"[翻页成功] 跳转至第 {page_number} 页")
            self.human_sleep(1.5, 2.5)

        except Exception as e:
            self.logger.warning(f"[翻页失败] 第 {page_number} 页: {e}")
            self.save_page_html("error_page.html")

    def run(self):
        self.driver.get("https://www.1688.com/")

        self.wait_for_login()
        #input("test")
        self.search()
        # 切换到搜索结果页
        handles = self.driver.window_handles
        self.driver.switch_to.window(handles[-1])

        # 如果不是从第一页开始，则尝试跳转
        if self.page_start != 1:
            self.multi_scroll_up_down()
            result = self.go_to_page(self.page_start)
            if result is not None:
                self.logger.warning(f"起始页 {self.page_start} 跳转失败，最大页码为 {result}，停止程序")
                return  # 或者 self.page_end = result 再继续处理部分页码

        for page in range(self.page_start, self.page_end + 1):
            if self.count - 2 >= self.max_items:
                self.logger.info(f"已抓取 {self.max_items} 个商品，停止抓取")
                break
            self.wait_if_paused()
            # 每页开始前检测停止和暂停
            if self.should_stop():
                self.logger.info("用户选择终止爬虫，提前结束运行。")
                break

            self.multi_scroll_up_down(8, 8, -800, 800)
            proceed = self.parse_page(page)
            if not proceed:
                if self.should_stop():
                    self.logger.info(f"用户选择终止爬虫，提前结束运行。")
                    break
                else:
                    self.logger.warning(f"第 {page} 页爬取失败，跳过该页。")
                    continue

            if page != self.page_end:
                # 翻页跳转下一页
                result = self.go_to_page(page + 1)
                if result is not None:
                    self.logger.warning(f"第 {page + 1} 页跳转失败，检测到最大页码为 {result}，提前结束")
                    break

        # 保存文件
        filename = f"{self.keyword}_{time.strftime('%Y%m%d-%H%M')}.xlsx"
        # 使用硬编码路径
        folder = "excel"  # 直接使用硬编码的路径
        if not os.path.exists(folder):
            os.makedirs(folder)
        # 构建完整的文件路径
        filepath = os.path.join(folder, filename)
        self.excel.save(filepath)
        self.logger.info(f"文件已保存: {filename}")
        self.clean_image_folder()
        # 获取当前窗口的句柄
        current_window_handle = self.driver.current_window_handle

        # 关闭当前窗口
        self.driver.close()

        # 获取所有标签页的句柄
        all_handles = self.driver.window_handles

        # 切换到并关闭剩余的标签页
        for handle in all_handles:
            if handle != current_window_handle:
                self.driver.switch_to.window(handle)
                self.driver.close()  # 关闭剩余的标签页
        self.driver.quit()


if __name__ == "__main__":
    keyword = "山东黄桃"#input("输入搜索关键词：")
    start_page =1 # int(input("起始页码："))
    end_page = 2 #int(input("终止页码："))
    max_items = 5 #int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"

    scraper = Ali_1688Scraper(keyword, start_page, end_page,insert_image, max_items=max_items)
    scraper.run()
