# -*- coding: utf-8 -*-
from encodings.punycode import T, selective_find
import time
import random
import json
import re
import requests
import os
import shutil
from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError
from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from pyquery import PyQuery as pq
from openpyxl import Workbook
from Spider.BaseScraper import BaseScraper

from Spider.AntiScrapingException import CaptchaHandler
from path_utils import static_image_path ,static_hash_path , static_excel_path


class JDScraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, insert_image = True, max_items=100, pierce_span=[0,0] ):
        super().__init__(headless=True, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.pierce_span = pierce_span
        self.all_warelists = []
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self.insert_image = insert_image
        self._setup_excel()
        self.hash_set = set()
        #self.lock = threading.Lock()  # 用于同步的锁
        #self.anti_spider_triggered = False
        self.captcha_handler = CaptchaHandler(self.driver, self.logger)
        self.captcha_handler.JDslider1()  # 调用风控
        self.captcha_handler.JDslider2()

    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def save_to_excel(self, results, filename='results.xlsx'):
        # 创建新的 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = '商品信息'

        # 初始化表头
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']

        # 写入表头
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)

        # 设置图片插入列宽
        img_col_letter = 'L'  # 图片插入列是第L列
        ws.column_dimensions[img_col_letter].width = 15

        # 填充商品数据
        for i, item in enumerate(results, start=2):  # 从第二行开始插入数据
            # 填充商品信息
            ws.cell(row=i, column=1, value=i - 1)  # 商品编号（Num）
            ws.cell(row=i, column=2, value=item['title'])  # 商品标题（Title）
            ws.cell(row=i, column=3, value=item['price'])  # 商品价格（Price）
            ws.cell(row=i, column=4, value=item['deal_num'])  # 成交量（Deal）
            ws.cell(row=i, column=5, value=item['location'])  # 商品所在地（Location）
            ws.cell(row=i, column=6, value=item['shop_name'])  # 店铺名称（Shop）
            ws.cell(row=i, column=7, value=item['freeFreight'])  # 是否包邮（IsPostFree）
            ws.cell(row=i, column=8, value=item['item_url'])  # 商品链接（Title_URL）
            ws.cell(row=i, column=9, value=item['shop_url'])  # 店铺链接（Shop_URL）
            ws.cell(row=i, column=10, value=item['img_url'])  # 图片链接（Img_URL）
            ws.cell(row=i, column=11, value=item['comment'])  # 评论数量（Num_Com）

            # 插入商品图片（Image）
            if self.insert_image:
                img_path = item.get('img_path')  # 商品图片路径
                if img_path and os.path.exists(img_path):
                    try:
                        # 使用 Pillow 打开图片并重新编码为 JPEG（修复部分 JPEG 插入失败问题）
                        with Image.open(img_path) as img:
                            rgb_img = img.convert("RGB")
                            rgb_img.save(img_path, format='JPEG')  # 覆盖原图

                        # 插入图片到 Excel
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 80
                        excel_img.height = 80
                        ws.row_dimensions[i].height = 60  # 设置行高
                        excel_img.anchor = f'{img_col_letter}{i}'  # 图片插入到指定单元格
                        ws.add_image(excel_img)
                    except Exception as e:
                        self.logger.error(f"[!] 图片插入失败: {img_path} | 错误: {e}")
                else:
                    self.logger.warning(f"[!] 图片路径无效或文件不存在: {img_path}")

        # 保存 Excel 文件
        wb.save(filename)
        self.logger.info(f"[√] 数据已保存到 Excel：{filename}")

    '''
    def save_cookies(self, path="JD_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="JD_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):#没做
        print("正在尝试登录JD...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)
            input("如出现滑块，请手动完成验证后按 Enter 继续...")
            self.save_cookies()
        except Exception as e:
            self.logger.warning(print("登录失败:", e))
    '''

    def search(self):
        try:
            self.logger.info("尝试定位搜索框和按钮...")

            search_box = None
            search_btn = None

            # ========= 第一优先：新版 JD（最稳 XPath） =========
            try:
                search_box = self.wait.until(
                    EC.element_to_be_clickable((
                        By.XPATH, '//input[@type="text" and @aria-label="搜索"]'
                    ))
                )

                search_btn = self.wait.until(
                    EC.element_to_be_clickable((
                        By.XPATH, '//button[.//div[text()="搜索"] or text()="搜索"]'
                    ))
                )

                self.logger.info("使用【新版 XPath】定位成功。")

            except Exception as e1:
                self.logger.warning(f"新版 XPath 失败: {e1}，尝试 class 定位...")

                # ========= 第二优先：新版 class（次稳） =========
                try:
                    search_box = self.wait.until(
                        EC.element_to_be_clickable((
                            By.CSS_SELECTOR, 'input.jd_pc_search_bar_react_search_input'
                        ))
                    )

                    search_btn = self.wait.until(
                        EC.element_to_be_clickable((
                            By.CSS_SELECTOR, 'button.jd_pc_search_bar_react_search_btn'
                        ))
                    )

                    self.logger.info("使用【新版 class】定位成功。")

                except Exception as e2:
                    self.logger.warning(f"class 定位失败: {e2}，尝试旧版...")

                    # ========= 第三优先：旧版 =========
                    search_box = self.wait.until(
                        EC.presence_of_element_located((By.ID, 'key'))
                    )

                    search_btn = self.wait.until(
                        EC.element_to_be_clickable((By.CLASS_NAME, 'button'))
                    )

                    self.logger.info("使用【旧版】定位成功。")

            # ========= 执行搜索 =========
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()

            self.logger.info(f"已点击搜索按钮，关键词：{self.keyword}")

            # ========= 风控处理 =========
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())

        except Exception as e:
            self.logger.error(f"搜索操作失败: {e}")

    def multi_scroll_up_down(self, up_num=1,down_num=8 , up_step=-200, down_step=800):
        '''包含JD风控检测模拟人类在页面上向下滚动加载内容，然后向上滚动查看已加载的内容，重复多次'''
        # 向下滚动
        for _ in range(down_num):
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())
            self.scroll_step_down(down_step)
    
        # 向上滚动
        for _ in range(up_num):
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())
            self.scroll_step_down(up_step)

        # 模拟人工休眠，防止过快的滚动
        time.sleep(random.uniform(1.0, 1.5))


    def go_to_page(self, page_number):
        #翻页前先滚动到页面底部，确保分页组件加载出来，翻页后包含风控检测
        self.RiskPause(self.captcha_handler.JDslider1())
        self.RiskPause(self.captcha_handler.JDslider2())

        try:
            # 定位输入框
            input_elem = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//div[contains(@class, "_pagination_toPageNum_")]//input')
            ))
            input_elem.clear()
            input_elem.send_keys(str(page_number))
            input_elem.send_keys(Keys.ENTER)  # 使用 Enter 键触发跳转

            # 等待页面跳转到目标页
            self.wait.until(EC.text_to_be_present_in_element(
                (By.XPATH, '//div[contains(@class, "_pagination_item_") and contains(@class, "_active_")]'),
                str(page_number)
            ))

            self.human_sleep(2, 3)
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())
        except Exception as e:
            self.logger.warning(f"跳转到第 {page_number} 页失败: {e}")
            self.save_page_html("error_page.html")

    def get_more(self, url):
        post_text = ''
        num_com = 0
        main_window = self.driver.current_window_handle

        try:
            # 提取商品ID
            item_id_match = re.search(r'(\d+)', url)
            if not item_id_match:
                self.logger.warning("无法提取商品ID，跳过")
                return '', 0
            item_id = item_id_match.group(1)

            # 定位商品卡片的可点击区域（新版京东结构）
            # 优先点击 a[href*="item.jd.com"]
            card_xpath = f'//div[@data-sku="{item_id}"]//a[contains(@href, "item.jd.com")]'
            try:
                card_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, card_xpath)))
            except Exception:
                # 兼容旧版 fallback
                card_xpath = f'//div[@data-sku="{item_id}"]//div[contains(@class, "_card_")]'
                card_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, card_xpath)))

            # 模拟点击商品卡片
            self.logger.debug(f"即将点击商品卡片：{item_id}")
            ActionChains(self.driver).move_to_element(card_element).click().perform()
            self.human_sleep(1, 2)

            # 等待新标签页打开
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return '', 0

            # 切换到新标签页
            new_tabs = [h for h in self.driver.window_handles if h != main_window]
            if not new_tabs:
                self.logger.warning("找不到新标签页句柄")
                self.save_page_html("error_page.html")
                return '', 0

            new_tab = new_tabs[-1]
            self.driver.switch_to.window(new_tab)
            self.human_sleep(1, 2)

            # 检查风控滑块
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())
            self.wait_if_paused()
            if self.should_stop():
                self.logger.info("用户主动终止爬虫，跳过当前评论数解析")
                self.stop()
                return '', 0

            # 获取包邮信息（新版 JD）
            try:
                elems = self.driver.find_elements(
                    By.XPATH,
                    '//div[contains(@class,"logistics-tips-wrapper")]//div[contains(@class,"tips-trigger")]'
                )

                if elems:
                    text = elems[0].text.strip()
                    post_text = "包邮" if "包邮" in text else ""
                else:
                    post_text = ""

            except:
                post_text = ""

            # 获取店铺链接（新版 JD）
            try:
                shop_elem = self.wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//a[contains(@href,"mall.jd.com/index-")]')
                ))

                shop_href = shop_elem.get_attribute('href')

                if shop_href:
                    if shop_href.startswith('//'):
                        shop_url = 'https:' + shop_href
                    else:
                        shop_url = shop_href
                else:
                    shop_url = ''

            except Exception:
                shop_url = ''

            # 获取评论数量（新版京东页面结构不同）
            try:
                # 尝试新版评论模块标题
                comment_elem = self.wait.until(EC.presence_of_element_located((
                    By.XPATH,
                    '//div[contains(@class, "comment") and (contains(text(), "评价") or contains(text(), "评论"))]'
                )))
                comment_text = comment_elem.text.strip()
                match = re.search(r'(\d+[万+]?)', comment_text)
                num_com = match.group(1) if match else 0
            except Exception:
                num_com = 0

        except Exception as e:
            self.logger.warning(f"详情失败: {e}")

        finally:
            # 关闭新标签页并切回主窗口
            self.human_sleep(1, 2)
            try:
                if self.driver.current_window_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    self.RiskPause(self.captcha_handler.JDslider1())
                    self.RiskPause(self.captcha_handler.JDslider2())
            except Exception as e:
                self.logger.warning(f"关闭标签页或切换回主窗口失败: {e}")
                self.save_page_html("error_page.html")

        return shop_url, post_text, num_com

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.jd.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(f"图片请求失败: {url}")
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(f"图片解码失败（格式可能不受支持，如webp）: {url}")
                return None
            except Exception as e:
                self.logger.error(f"图片处理异常: {e} | URL: {url}")
                return None

            # 创建文件夹（硬编码路径）
            folder = "images/JD"  # 使用硬编码路径
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 文件名硬编码处理
            file_name = f"{index:04d}.jpg"  # 保持文件名格式不变
            file_path = os.path.join(folder, file_name)  # 拼接路径

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(f"图片下载异常: {e} | URL: {url}")
            return None

    def download_all_images(self):
        """
        批量下载 self.all_warelists 中的图片并保存到本地。
        并在每个商品字典中添加 'image_path' 字段。
        """
        for index, product_info in enumerate(self.all_warelists):
            img_url = product_info.get('img_url', 'N/A')

            if img_url != 'N/A':  # 如果图片链接有效
                image_path = self.download_image(img_url, index + 1)  # 下载并保存图片
                if image_path:
                    product_info['img_path'] = image_path  # 更新商品字典，保存图片的本地路径
                else:
                    product_info['img_path'] = None  # 下载失败则设为 None
            else:
                product_info['img_path'] = None  # 如果没有图片链接，也设为 None

        self.logger.info(f"已成功下载并保存 {len(self.all_warelists)} 张商品图片")


    def clean_image_folder(self):
        # 获取图片文件夹路径并清空（硬编码路径）
        folder = "images/JD"  # 使用硬编码路径
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)  # 删除文件或快捷方式
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)  # 删除子文件夹
                except Exception as e:
                    self.logger.warning(f"[!] 删除缓存文件失败: {file_path} | 错误: {e}")
        else:
            self.logger.error(f"[!] 文件夹不存在或不是目录：{folder}")

    def JSlistener(self, driver):
        """注入 JavaScript 脚本来拦截 fetch 和 XHR 请求，并获取商品数据"""
        script = """
        (function() {
            // 检查 sessionStorage 中是否已经设置过拦截器标识
            if (sessionStorage.getItem('jd_interceptor_installed') === 'true') return;
        
            // 设置标识符，确保拦截器只注入一次
            sessionStorage.setItem('jd_interceptor_installed', 'true');

            // 缓存数组
            window._jd_responses = [];

            // ====== fetch 拦截 ======
            const origFetch = window.fetch;
            window.fetch = async (...args) => {
                const response = await origFetch(...args);
                try {
                    const url = args[0];
                    if (typeof url === 'string' && url.includes('api.m.jd.com')) {
                        const cloned = response.clone();
                        cloned.json().then(data => {
                            if (data) {
                                // 不限制数量，直接将数据加入缓存数组
                                window._jd_responses.push(data);
                                console.log('[拦截到 fetch 搜索响应]', data);

                                // 保存到 LocalStorage
                                let savedData = JSON.parse(localStorage.getItem('jd_search_data')) || [];
                                savedData.push(data);
                                localStorage.setItem('jd_search_data', JSON.stringify(savedData));  // 将数据保存到 LocalStorage
                            }
                        }).catch(err => console.warn('[fetch JSON 解析失败]', err));
                    }
                } catch (err) {
                    console.warn('[fetch 拦截异常]', err);
                }
                return response;
            };

            // ====== XHR 拦截 ======
            const origXHRSend = XMLHttpRequest.prototype.send;
            XMLHttpRequest.prototype.send = function(body) {
                this.addEventListener('load', function() {
                    try {
                        const url = this.responseURL || '';
                        if (url.includes('api.m.jd.com') && this.responseType === '' && this.responseText) {
                            const data = JSON.parse(this.responseText);
                            if (data) {
                                // 不限制数量，直接将数据加入缓存数组
                                window._jd_responses.push(data);
                                console.log('[拦截到 XHR 搜索响应]', data);

                                // 保存到 LocalStorage
                                let savedData = JSON.parse(localStorage.getItem('jd_search_data')) || [];
                                savedData.push(data);
                                localStorage.setItem('jd_search_data', JSON.stringify(savedData));  // 将数据保存到 LocalStorage
                            }
                        }
                    } catch (e) {
                        console.warn('[XHR 解析失败]', e);
                    }
                });
                origXHRSend.apply(this, arguments);
            };
        })();
        """    
        # 执行 JavaScript 脚本来注入拦截器
        driver.execute_script(script)

    def clear_jd_search_responses(self, driver):
        """仅清空 LocalStorage 中的具体数据，避免重用旧数据，而不清空 sessionStorage 中的标识符。"""
        driver.execute_script("""
        // 仅清空 LocalStorage 中的抓取数据，不清空 sessionStorage 中的标识符
        localStorage.removeItem('jd_search_data');  // 清空抓取的数据
        console.log('LocalStorage 中的 jd_search_data 已清空');
        """)

    def parse_page(self, page_number,count_number = 0):
        count_number = len(self.all_warelists)
        #暂停检测
        self.wait_if_paused()
        #终止检测
        if self.should_stop():
            self.logger.info("检测到终止命令")
            return 'finish'  # 返回 'finsih' 表示抓取终止
        try:
            # 注入拦截器
            self.JSlistener(self.driver)
            self.human_sleep(1,2)
            # 滚动页面以加载数据
            self.multi_scroll_up_down(0, 8, 0, 600)  # 前往目标页面
            self.go_to_page(page_number)
            self.clear_jd_search_responses(self.driver)  # 清空缓存，避免旧数据干扰

            self.human_sleep(1, 2)
            self.multi_scroll_up_down(8, 8, -800, 800)  # 加载完整数据

            # 获取 LocalStorage 中保存的数据
            saved_data = self.driver.execute_script("""
            return JSON.parse(localStorage.getItem('jd_search_data'));
            """)

            if not saved_data:
                self.logger.warning(f"第{page_number}页没有找到任何商品数据")
                return 'no_data'  # 返回 'no_data' 表示没有数据，外层循环可以停止

            # 正向遍历 saved_data 列表
            for item in saved_data:
                if isinstance(item, dict) and 'data' in item:
                    data = item['data']
                    if isinstance(data, dict) and 'wareList' in data:
                        warelist = data['wareList']
                        if warelist:  # 如果 wareList 不为空
                            for product in warelist:
                                product_info = {
                                    'title': product.get('wareName', 'N/A'),
                                    'price': product.get('realPrice', 'N/A'),
                                    'deal_num': product.get('totalSales', 'N/A'),
                                    'location': "",
                                    'shop_name': product.get('shopName', 'N/A'),
                                    'freeFreight': product.get('freeFreight', 'N/A'),
                                    'shop_id': product.get('shopId', 'N/A'),
                                    'skuId': product.get('skuId', 'N/A'),
                                    'comment': product.get('comment', 'N/A'),
                                    'img_url': product.get('imageurl', 'N/A')
                                }

                                # 清理标题字段，去除 HTML 标签和解码 HTML 实体字符
                                cleaned_title = self.remove_html_tags(product_info['title'])  # 去除HTML标签
                                cleaned_title = self.decode_html_entities(cleaned_title)  # 解码HTML实体字符
                                product_info['title'] = cleaned_title

                                # 拼接完整的 URL
                                product_info['item_url'] = f"https://item.jd.com/{product_info['skuId']}.html"
                                product_info['shop_url'] = f"https://mall.jd.com/index-{product_info['shop_id']}.html"
                                product_info['img_url'] = f"https://img10.360buyimg.com/n2/s480x480_{product_info['img_url']}"

                                # ==== 价格过滤 ==== 
                                min_p, max_p = self.pierce_span

                                # 防止用户输入反区间 [100, 50]
                                if min_p > max_p:
                                    min_p, max_p = max_p, min_p

                                # 只有当区间不是 [0,0] 时才进行过滤
                                if not (min_p == 0 and max_p == 0):
                                    try:
                                        # 获取商品价格（从 `product_info['price']` 获取）
                                        price_value = float(product_info['price']) if product_info['price'] != 'N/A' else 0
                                    except:
                                        self.logger.warning(f"无法解析价格 {product_info['price']}")
                                        continue  # 无法解析价格 → 跳过

                                    # 不在闭区间 → 跳过该商品
                                    if not (min_p <= price_value <= max_p):
                                        self.logger.info(f"[价格过滤] {price_value} 不在区间 [{min_p}, {max_p}] 内，跳过")
                                        continue

                                # ==== 哈希去重 ====
                                # 使用 compute_hash 方法生成商品的哈希值
                                item_dict = {'title': product_info['title'], 'price': product_info['price']}
                                features_url = product_info['item_url']
                                product_hash = self.compute_hash(item_dict, 'JD', features_url)

                                # 如果商品哈希值已经存在，跳过
                                if product_hash in self.hash_set:
                                    self.logger.info(f"[跳过] 已存在商品: {product_info['title']}")
                                    continue

                                # 添加哈希值到集合
                                self.hash_set.add(product_hash)

                                # 将商品信息添加到当前页商品列表
                                self.all_warelists.append(product_info)
                                count_number += 1
                                self.count += 1
                                self.logger.info(f"提取到商品数量: {count_number}")

                                # 如果已经抓取到最大数量的商品，停止处理
                                if count_number >= self.max_items:
                                    self.logger.info(f"已抓取 {self.max_items} 条商品数据，停止抓取此页数据...")
                                    return 'max_reached'  # 返回 'max_reached' 表示达到最大数量

            return 'continue'  # 返回 'continue' 表示数据抓取成功并继续抓取下一页
    
        except Exception as e:
            # 记录详细的错误信息
            self.logger.error(f"提取商品信息失败: 第{page_number}页, 错误: {str(e)}")
            return 'error'  # 返回 'error' 表示抓取失败，外层循环可以决定是否继续


    def turn_page(self, page_number):
        try:
            # 定位输入框（新版）
            input_box = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//div[contains(@class, "_pagination_toPageNum_")]//input')
            ))
            input_box.clear()
            input_box.send_keys(str(page_number))
            input_box.send_keys(Keys.ENTER)  # 按下 Enter 触发跳页

            # 等待当前页码变更为目标页
            self.wait.until(EC.text_to_be_present_in_element(
                (By.XPATH, '//div[contains(@class, "_pagination_item_") and contains(@class, "_active_")]'),
                str(page_number)
            ))

            self.human_sleep(2, 3)
            self.RiskPause(self.captcha_handler.JDslider1())
            self.RiskPause(self.captcha_handler.JDslider2())

        except Exception as e:
            self.logger.error(f"翻页失败，目标页码 {page_number}：{e}")
            self.save_page_html("error_page.html")



    def run(self):
        try:
            # 打开京东主页
            self.driver.get("https://www.jd.com/")
        
            # 等待前端登录完成（你可以根据需要实现登录检查）
            self.wait_for_login()  # 如果需要等待登录完成，取消注释
        
            # 通过输入来确认开始
            #input("Press Enter to start parsing...")
            self.search()
        
            # 初始化抓取的商品信息列表
            self.all_warelists = []

            # 循环遍历指定的页面范围
            for page_number in range(self.page_start, self.page_end + 1):

                #暂停检测
                self.wait_if_paused()
                #终止检测
                if self.should_stop():
                    self.logger.info("检测到终止命令")
                    break

                self.logger.info(f"开始抓取第 {page_number} 页的数据...")

                # 调用 parse_page 进行数据抓取
                result = self.parse_page(page_number)
                # 根据返回的结果决定是否继续抓取
                if result == 'max_reached':
                    self.logger.info(f"已抓取 {self.max_items} 条商品数据，停止抓取...")
                    break  # 达到最大数量，停止抓取
                elif result == 'no_data':
                    self.logger.info(f"第 {page_number} 页没有数据，跳过...")
                    break  # 没有数据终止
                elif result == 'error':
                    self.logger.error(f"第 {page_number} 页抓取失败，跳过...")
                elif result == 'finish':
                    break  # 发生错误终止
                self.human_sleep(1, 2)

            # 打印或处理抓取的商品数据
            self.logger.info(f"共抓取到 {len(self.all_warelists)} 条商品数据")

            data = self.all_warelists[:self.max_items]  # 只保留前 max_items 条数据
            if data:
                # 使用硬编码路径
                filename = f"{self.keyword}_{time.strftime('%Y%m%d_%H%M')}.xlsx"
                folder = "excel"  # 硬编码路径
                if not os.path.exists(folder):
                    os.makedirs(folder)  # 如果目录不存在，则创建

                # 生成完整的文件路径
                filepath = os.path.join(folder, filename)
                self.download_all_images()  # 批量下载并更新图片路径
                self.save_to_excel(data, filepath)  # 保存数据到 Excel
            self.clean_image_folder()  # 清理图片缓存


        except Exception as e:
            # 记录详细的错误信息
            self.logger.error(f"发生错误：{e}")

        self.driver.close()
        self.driver.quit()

if __name__ == "__main__":
    keyword = "豆浆"#input("输入搜索关键词：")
    start_page =1 # int(input("起始页码："))
    end_page = 5 #int(input("终止页码："))
    max_items = 20 #int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"

    scraper = JDScraper(keyword, start_page, end_page,insert_image, max_items=max_items)
    scraper.run()
