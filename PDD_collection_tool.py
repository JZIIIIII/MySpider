# -*- coding: utf-8 -*-

import time
import random
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError
from tkinter import filedialog

from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from pyquery import PyQuery as pq
from openpyxl import Workbook
from BaseScraper  import BaseScraper



class PDDSpider(BaseScraper):
    def __init__(self, keyword, start_page, end_page, max_items=100, insert_image=True):
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.insert_image = insert_image
        self.count = 2
        self.logger = self._init_logger()
        self.driver = self.init_driver()
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self._setup_excel()


    def _setup_excel(self):
        # 根据修改后的 save_to_excel 调整表头顺序
        headers = ['Num', 'Title', 'Price', 'Deal', 'shop_url', 'CommentNum', 'ShopName', 'Postage', 'Tags', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def human_sleep(self, min_time=1, max_time=2):
        time.sleep(random.uniform(min_time, max_time))

    def scroll_step_down(self, base_step=800):
        """模拟人类向下较大幅度滑动"""
        step = random.randint(base_step - 300, base_step + 300)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.8, 1.5))  # 适当增加等待时间，保证加载

    def simulate_click(self, element):
        try:
            actions = ActionChains(self.driver)
            actions.move_to_element(element).click().perform()
            self.logger.info(print("模拟鼠标点击成功"))
            return True
        except Exception as e:
            self.logger.error(print("模拟鼠标点击失败:", e))
            self.save_page_html("error_page.html")

            return False

    def click_fake_search_box(self, timeout=10):
        wait = WebDriverWait(self.driver, timeout)
        try:
            time.sleep(2)  # 等待页面加载

            # 等待目标元素可点击
            element = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "_18v23kPu")))

            # 调用封装好的模拟点击函数
            return self.simulate_click(element)

        except Exception as e:
            self.logger.error(print("搜索框模拟点击失败:", e))
            self.save_page_html("error_page.html")

            return False

    def search(self):
        try:
            # 先尝试定位真实搜索框
            print("尝试定位真实搜索框...")
            real_search_box = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@type='search' and contains(@class, '_2bfwu6WT')]")
            ))
        except Exception:
            print("请尝试点击首页搜索框进入搜索页面...")
            return  # 找不到搜索框，结束函数，等待用户点击假搜索框进入搜索页

        try:
            # 定位搜索按钮
            search_btn = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'RuSDrtii') and text()='搜索']")
            ))
            # 输入关键词
            print(f"输入关键词: {self.keyword}")
            real_search_box.clear()
            real_search_box.send_keys(self.keyword)

            # 等待 1~2 秒，模拟人操作
            self.human_sleep(1, 2)

            # 用你的 simulate_click 函数点击搜索按钮
            self.simulate_click(search_btn)

            # 等待 1~2 秒，等待搜索结果加载
            self.human_sleep(1, 2)

            print("搜索成功")

        except Exception as e:
            self.log.err(print("搜索失败:", repr(e)))
            self.save_page_html("error_page.html")




    def parse_all_showcases(self, max_items=10):
        results = []
        seen_titles = set()
        processed_count = 0

        for _ in range(30):  # 最多下滑 30 次
            items = self.driver.find_elements(By.CSS_SELECTOR, 'div.rjNMXsUm._1unt3Js-')
            for index in range(len(items)):
                if len(results) >= max_items:
                    break

                try:
                    items = self.driver.find_elements(By.CSS_SELECTOR, 'div.rjNMXsUm._1unt3Js-')  # 每次都重新获取（防止 stale）
                    if index >= len(items):
                        continue

                    elem = items[index]
                    item_html = elem.get_attribute('outerHTML')
                    item_doc = pq(item_html)

                    # 提取标题并去重
                    title = item_doc('div._3ANzdjkc').text().strip()
                    if not title or title in seen_titles:
                        continue
                    seen_titles.add(title)

                    # 提取价格（兼容多种价格结构）
                    price = ''
                    price_dec_elem = item_doc('span._2aP8LGPL')
                    if price_dec_elem:
                        price_int = price_dec_elem.prev().text()
                        price_dec = price_dec_elem.text()
                        price = f"{price_int}{price_dec}"
                    elif not price:
                        price_elem = item_doc('span._3_U04GgA')
                        if price_elem:
                            coupon_tag = price_elem('span._2nKWnaqa')
                            if coupon_tag:
                                price = price_elem('span._3f_Cp5GQ + span').text()
                            else:
                                price = price_elem('span._3f_Cp5GQ + span').text()

                    # 提取标签
                    tag_list = [tag.text() for tag in item_doc('div._299OVZvt > div').items()]
                    tags_str = ' '.join(tag_list)

                    # 提取图片链接
                    img_elem = item_doc('div._1o7l_Qm- img')
                    img_url = img_elem.attr('src') or img_elem.attr('data-src') or img_elem.attr('data-lazy')

                    # 提取销量
                    deal_num_elem = item_doc('div[style*="width: 255px;"] span')
                    deal_num_text = deal_num_elem.text().strip()
                    deal_num = ''.join([ch for ch in deal_num_text if ch.isdigit()])
                    if not deal_num:
                        deal_num = 0

                    # 进入详情页或悬浮店铺卡片，获取更多信息
                    comment_count, shop_name, postage_info, shop_url = self.get_more(elem)
                    if comment_count is None:
                        comment_count = 0

                    # 下载图片，保存本地路径（新增）
                    img_path = None
                    if self.insert_image and img_url:
                        img_path = self.download_image(img_url, title)

                    # 汇总结果
                    result = {
                        'title': title,
                        'price': price,
                        'deal_num': deal_num,
                        'shop_url': shop_url,
                        'comment_count': comment_count,
                        'shop_name': shop_name,
                        'postage_info': postage_info,
                        'tags': tags_str,
                        'img_url': img_url,
                        'img_path': img_path  # 本地图片路径，供 Excel 插图使用
                    }
                    results.append(result)

                    processed_count += 1
                    if processed_count % 2 == 0:
                        self.scroll_step_down(base_step=800)
                        time.sleep(1)

                except Exception as e:
                    self.logger.warning(print(f"[!] 处理商品失败: {e}"))
                    continue

            if len(results) >= max_items:
                break

            # 向下滚动加载更多
            self.scroll_step_down(base_step=1200)
            time.sleep(1)

        print(f"\n共提取到 {len(results)} 个商品（上限：{max_items}）")

        self.clear_image_cache('images')  # 清空 image 文件夹缓存

        return results

    def save_to_excel(self, results, filename='results.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = '商品信息'

        # 表头
        headers = ['标题', '价格', '成交量', '店铺连接', '评论数量', '店铺名称', '包邮信息', '标签', '图片']
        ws.append(headers)

        # 设置图片插入列列宽（第9列）
        img_col_letter = 'I'
        ws.column_dimensions[img_col_letter].width = 15

        for i, item in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=item['title'])
            ws.cell(row=i, column=2, value=item['price'])
            ws.cell(row=i, column=3, value=item['deal_num'])
            ws.cell(row=i, column=4, value=str(item['shop_url']))
            ws.cell(row=i, column=5, value=item['comment_count'])
            ws.cell(row=i, column=6, value=item['shop_name'])
            ws.cell(row=i, column=7, value=item['postage_info'])
            ws.cell(row=i, column=8, value=item['tags'])

            # 如果开启插图功能
            if self.insert_image:
                img_path = item.get('img_path')
                if img_path and os.path.exists(img_path):
                    try:
                        # 尝试用 Pillow 打开并重新编码为 JPEG（修复部分 JPEG 插入失败问题）
                        with Image.open(img_path) as img:
                            rgb_img = img.convert("RGB")
                            rgb_img.save(img_path, format='JPEG')  # 覆盖原图
                    
                        # 插入图片
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 80
                        excel_img.height = 80
                        ws.row_dimensions[i].height = 60
                        excel_img.anchor = f'{img_col_letter}{i}'
                        ws.add_image(excel_img)
                    except Exception as e:
                        self.logger.error(print(f"[!] 图片插入失败: {img_path} | 错误: {e}"))
                else:
                    self.logger.warning(print(f"[!] 图片路径无效或文件不存在: {img_path}"))

        # 保存 Excel 文件
        wb.save(filename)
        print(f"[√] 数据已保存到 Excel：{filename}")

    def webp_to_png(self, webp_bytes, save_path):
        try:
            img = Image.open(BytesIO(webp_bytes))
            if img.format != 'WEBP':
                self.logger.warning(print(f"警告：不是WEBP格式，实际是 {img.format}"))
            img = img.convert("RGBA")  # 保留透明度
            img.save(save_path, format="PNG")
            self.logger.info(print(f"WEBP图片转换PNG并保存成功: {save_path}"))
            return save_path
        except Exception as e:
            self.logger.warning(print(f"WEBP转PNG失败: {e}"))
            return None

    def jpeg_to_jpg(self, jpeg_bytes, save_path):
        try:
            img = Image.open(BytesIO(jpeg_bytes))
        
            # 检查格式是否为 JPEG
            if img.format != 'JPEG':
                self.logger.warning(print(f"警告：不是JPEG格式，实际是 {img.format}"))
                return None

            # 强制保存为 JPG 格式
            img = img.convert("RGB")
            img.save(save_path, format="JPEG")
            self.logger.info(print(f"JPEG图片转换JPG并保存成功: {save_path}"))
            return save_path
        except Exception as e:
            self.logger.warning(print(f"JPEG转JPG失败: {e}"))
            return None

    def download_image(self, img_url, title, save_dir='images'):
        # 确保保存目录存在
        os.makedirs(save_dir, exist_ok=True)

        # 清理标题中的特殊字符，避免路径错误
        filename = re.sub(r'[\\/:*?"<>|]', '_', title[:20].strip().replace(' ', '_')) + ".jpg"
        path = os.path.join(save_dir, filename)

        # 处理相对路径
        if img_url.startswith("//"):
            img_url = "https:" + img_url
        elif img_url.startswith("/"):
            img_url = "https://img.pddpic.com" + img_url  # 拼多多图片域名

        img_url = img_url.split('?')[0]  # 去掉查询字符串

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "Referer": "https://mobile.pinduoduo.com/",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        }

        try:
            resp = requests.get(img_url, headers=headers, timeout=10)
            if resp.status_code != 200:
                self.logger.error(print(f"图片下载失败，状态码: {resp.status_code} | URL: {img_url}"))
                return None

            try:
                # 打开图片
                img = Image.open(BytesIO(resp.content))

                # 如果图片是 WebP 格式，转换为 PNG 格式
                if img.format == 'WEBP':
                    print(f"[!] 图片为WEBP，转换为PNG: {img_url}")
                    png_path = os.path.splitext(path)[0] + ".png"
                    result = self.webp_to_png(resp.content, png_path)
                    return result

                # 转换为 RGB 格式并保存为 JPEG
                img = img.convert('RGB')
                img.save(path, format='JPEG')
                print(f"图片已保存: {path}")
                return path

            except UnidentifiedImageError:
                self.logger.error(print(f"[!] 图片格式不支持或路径无效: {img_url}"))
                return None
            except Exception as e:
                self.logger.error(print(f"[!] 图片处理异常: {e} | URL: {img_url}"))
                return None

        except Exception as e:
            self.logger.error(print(f"[!] 图片下载异常: {e} | URL: {img_url}"))
            return None

    def clean_url(self, url):
        """
        使用正则表达式清理 URL，保留 goods_id 参数，去掉其他多余的参数。
        """
        try:
            # 使用正则匹配并提取 base_url 和 goods_id 参数
            match = re.search(r"(https?://[^\?]+)(\?[^#]*)", url)
            if match:
                base_url = match.group(1)  # 基础 URL
                query_string = match.group(2)  # 查询部分

                # 解析查询参数
                query_params = parse_qs(query_string[1:])  # 去掉 '?'
                
                # 只保留 'goods_id' 参数
                cleaned_params = {key: value for key, value in query_params.items() if key == 'goods_id'}
                
                # 构造新的查询字符串
                cleaned_query = urlencode(cleaned_params, doseq=True)
                
                # 如果有有效的 cleaned_query，则重建新的 URL
                if cleaned_query:
                    cleaned_url = f"{base_url}?{cleaned_query}"
                else:
                    cleaned_url = base_url  # 如果没有有效的参数，返回没有参数的 URL
                
                return cleaned_url
            else:
                return url  # 如果没有匹配，返回原始 URL
        except Exception as e:
            self.logger.error(print(f"Error cleaning URL: {e}"))
            return url

    def get_more(self, element):
        comment_count = 0
        shop_name = ''
        postage_info = ''
        shop_url = ''  # 新增用于存储店铺URL

        try:
            # 先滚动到元素中间，避免被遮挡
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.5)

            # 等待遮挡元素消失（如果知道遮挡的class，替换下）
            try:
                self.wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, 'div.浮动遮挡层的class')))
            except TimeoutException:
                pass  # 或打印日志

            # 直接用 JS 点击，绕过遮挡
            self.driver.execute_script("arguments[0].click();", element)
        
            # 获取当前页面 URL
            current_url = self.driver.current_url
            shop_url = self.clean_url(current_url)
        
            # 等待详情页关键元素加载
            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.F2MXl7Xc')))
            except Exception:
                self.logger.warning(print("[!] 无法加载评论数量元素，跳过"))
                comment_count = 0  # 默认评论数为0

            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.BAq4Lzv7')))
            except Exception:
                self.logger.warning(print("[!] 无法加载店铺名元素，跳过"))
                shop_name = "无店铺名称"  # 默认店铺名称

            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.RbQ7MTuU')))
            except Exception:
                self.logger.warning(print("[!] 无法加载包邮信息元素，跳过"))
                postage_info = "无包邮信息"  # 默认包邮信息

            # 获取评论数（如果没有评论，这一步会失败）
            try:
                comment_text_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.F2MXl7Xc'))
                )
                comment_text = comment_text_element.text
                match = re.search(r'\((\d+)\)', comment_text)
                if match:
                    comment_count = int(match.group(1))
            except Exception:
                comment_count = 0  # 如果没有评论或无法加载评论，设置为 0

            # 获取店铺名称
            try:
                shop_name_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.BAq4Lzv7'))
                )
                shop_name = shop_name_element.text.strip()
            except Exception as e:
                self.logger.warning(print(f"[!] 获取店铺名称失败: {repr(e)}"))
                shop_name = "无店铺名称"  # 设置默认值


            # 获取包邮状况
            try:
                postage_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.RbQ7MTuU span.KDFIGUNK'))
                )
                postage_info = ' '.join([e.text.strip() for e in postage_elements if e.text.strip()])
            except Exception as e:
                self.logger.warning(print(f"[!] 获取包邮信息失败: {repr(e)}"))
                postage_info = "无包邮信息"  # 设置默认值

        except Exception as e:
            self.logger.error(print(f"[!] 获取详情数据失败: {repr(e)}"))
            self.save_page_html("error_page.html")


        finally:
            self.driver.back()
            try:
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.rjNMXsUm._1unt3Js-')))
            except TimeoutException:
                time.sleep(1)
            self.human_sleep(1, 2)
        
        #print(shop_url)
        return comment_count, shop_name, postage_info, shop_url  # 返回店铺URL

    def clear_image_cache(self,folder='images'):
        """
        清空指定文件夹内所有文件和子文件夹，默认清空 image 文件夹
        """
        if os.path.exists(folder) and os.path.isdir(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    self.logger.warning(print(f"[!] 删除缓存文件失败: {file_path} | 错误: {e}"))
        else:
            self.logger.error(print(f"[!] 文件夹不存在或不是目录：{folder}"))

    def run(self):
        try:
            self.driver.get("https://mobile.pinduoduo.com/")
            self.log.info(print("登录拼多多成功"))
        except:
            self.logger.error(print("登录拼多多失败"))
        input("请登录拼多多并手动跳转到首页后按回车继续...")
        self.click_fake_search_box()
        self.search()
        data = self.parse_all_showcases(max_items=self.max_items)
    
        if data:
            filename = f"{self.keyword}_{time.strftime('%Y%m%d_%H%M')}.xlsx"
            self.save_to_excel(data, filename)
    
        self.driver.quit()


if __name__ == "__main__":
    kw = input("输入关键词：")
    sp = 1 #int(input("起始页码："))
    ep = 1 #int(input("结束页码："))
    mi = int(input("最大商品数："))
    show_img = input("是否插入图片 (y/n)：").strip().lower() == 'y'

    spider = PDDSpider(kw, sp, ep, mi, show_img)
    spider.run()
