<<<<<<< HEAD
﻿# -*- coding: utf-8 -*-

import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from BaseScraper import BaseScraper



class Ali_1688Scraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, max_items=100):
        #super().__init__(headless=None, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.logger = self._init_logger()
        self.driver = self.init_driver()
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self.insert_image = True
        self._setup_excel()
        #self.lock = threading.Lock()  # 用于同步的锁

    def human_sleep(self, min_time=0.5, max_time=1.5):
        time.sleep(random.uniform(min_time, max_time))

    def check_for_captcha(self):
        try:
            # 使用 self.driver 查找页面中特定的验证码元素
            captcha_element = self.driver.find_element_by_xpath("//div[@class='captcha']")
            if captcha_element:
                print("检测到验证码！暂停程序。")
                return True
        except Exception as e:
            # 如果没有找到验证码相关元素，返回 False
            return False




    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def save_cookies(self, path="taobao_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="taobao_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):
        print("正在尝试登录1688...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)
            input("如出现滑块，请手动完成验证后按 Enter 继续...")
            self.save_cookies()
        except Exception as e:
            self.logger.waring(print("登录失败:", e))

    def search(self):
        try:
            print("尝试用定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'alisearch-input')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'input-button')))
            self.check_for_captcha()
        except Exception as e:
            print("搜索框定位失败")
        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.check_for_captcha()
            self.human_sleep(2, 3)
        except Exception as e:
            self.logger.warning(print("搜索操作失败:", e))
            self.save_page_html("error_page.html")

    def go_to_page(self, page_number):
        try:
            # 定位输入框，清空并输入页码
            page_input = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'input.input-page')))
            page_input.clear()
            page_input.send_keys(str(page_number))
            self.human_sleep(0.5, 1.2)

            # 定位并点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'div.paging-to-page-button')))
            confirm_btn.click()
            self.check_for_captcha()
            self.human_sleep(2, 3)

        except Exception as e:
            self.logger.warning(f"翻页失败: {e}")
            self.save_page_html("error_page.html")


    def get_more(self, url):
        comment_count = 0
        location = ''
        main_window = self.driver.current_window_handle

        try:
            # 提取商品ID，防止href匹配失败
            item_id_match = re.search(r'/offer/(\d+)\.html', url)
            if not item_id_match:
                self.logger.warning("无法提取商品ID，跳过")
                return location, comment_count
            item_id = item_id_match.group(1)

            # 限制 href 必须包含 detail.1688.com
            title_xpath = f'//a[contains(@href, "/offer/{item_id}.html") and contains(@href, "detail.1688.com")]'
            title_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, title_xpath)))

            href = title_element.get_attribute('href')
            #print(f"即将点击链接：{href}")

            # 获取尺寸并设置更合适的偏移
            size = title_element.size
            width = size['width']
            height = size['height']
            offset_x = int(width * 0.1)
            offset_y = int(height * 0.3)

            # 执行偏移点击
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL) \
                   .move_to_element_with_offset(title_element, offset_x, offset_y) \
                   .click() \
                   .key_up(Keys.CONTROL) \
                   .perform()

            self.check_for_captcha()
            # 等待新标签页打开，最多等待5秒
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return location, comment_count

            # 切换到新标签页
            handles = self.driver.window_handles
            new_tabs = [h for h in handles if h != main_window]
            if not new_tabs:
                self.logger.warning("找不到新标签页句柄")
                self.save_page_html("error_page.html")
                return location, comment_count

            new_tab = new_tabs[-1]
            self.driver.switch_to.window(new_tab)
            self.check_for_captcha()
            self.human_sleep(2, 3)

            # 获取评论数
            try:
                element = self.driver.find_element(By.CSS_SELECTOR, 'span.brackets')
                comment_count = int(element.get_attribute("data-value"))
            except Exception as e:
                self.logger.warning(f"获取评论数失败，错误信息：{e}")
                comment_count = 0

            # 获取发货地
            try:
                location_element = self.driver.find_element(By.CSS_SELECTOR, 'span.location')
                location = location_element.text.strip()
            except Exception as e:
                self.logger.warning(f"获取发货地失败，错误信息：{e}")
                location = ""

        except Exception as e:
            self.logger.warning(f"标签页打开失败，错误信息：{e}")

        finally:
            # 关闭新标签页，切换回主窗口
            self.human_sleep(1, 2)
            try:
                current_handle = self.driver.current_window_handle
                if current_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    self.check_for_captcha()

            except Exception as e:
                self.logger.warning(f"关闭标签页或切换主窗口失败，错误信息：{e}")
                self.save_page_html("error_page.html")

        return location, comment_count

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.1688.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(print(f"图片请求失败: {url}"))
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(print(f"图片解码失败（格式可能不受支持，如webp）: {url}"))
                return None
            except Exception as e:
                self.logger.error(print(f"图片处理异常: {e} | URL: {url}"))
                return None

            # 创建文件夹
            folder = "images"
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 生成文件名，保存为 jpg
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(print(f"图片下载异常: {e} | URL: {url}"))
            return None

    def clean_image_folder(self):
        folder = "images"
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # 删除文件或快捷方式
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # 删除子文件夹

    ##滑动窗口
    def scroll_step_down(self, base_step=300):
        """模拟人类向下小段滑动"""
        step = random.randint(base_step - 100, base_step + 100)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.6, 1.2))
        #self.wait_manual_verification()


    def smart_scroll_until_loaded(self, min_new_items=4, max_scroll_attempts=10):
        """边滚动边检测新商品是否加载，最多尝试 max_scroll_attempts 次"""
        last_count = 0
        attempts = 0

    def multi_scroll_up_down(self, up_num=1,down_num=8 , up_step=-200, down_step=600):
        # 向下滚动
        for _ in range(down_num):
            self.scroll_step_down(down_step)
    
        # 向上滚动
        for _ in range(up_num):
            self.scroll_step_down(up_step)
 

        # 模拟人工休眠，防止过快的滚动
        time.sleep(random.uniform(1.0, 1.5))

    def parse_page(self, page_number):
        # 获取当前页面所有商品元素（Selenium WebElement）
        items_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a.search-offer-item')

        # 用 pyquery 解析页面HTML，拿到所有商品节点
        html = self.driver.page_source
        doc = pq(html)
        items_data = list(doc('a.search-offer-item').items())

        if len(items_elements) != len(items_data):
            self.logger.warning(f"警告：Selenium 找到商品数量 {len(items_elements)} 与 pyquery 找到数量 {len(items_data)} 不一致，可能页面未完全加载！")
            # 这里可以等待或重试，这里先简单返回False停止
            return False

        slide_counter = 0  # 下滑计数器

        for idx, (element, data) in enumerate(zip(items_elements, items_data)):
            if self.count - 2 >= self.max_items:
                print(f"已达到最大抓取数量：{self.max_items}，停止抓取。")
                return False

            # 滚动到当前商品，使其位于中央
            self.driver.execute_script("""
                var elem = arguments[0];
                var rect = elem.getBoundingClientRect();
                var windowHeight = window.innerHeight;
                var elementTop = rect.top;
                var elementBottom = rect.bottom;
                var scrollAmount = (elementTop + elementBottom) / 2 - windowHeight / 2;
                window.scrollBy(0, scrollAmount);
            """, element)

            # 解析图片URL
            img_tag = data.find('.offer-img-inner img') or data.find('img')
            if not img_tag:
                img_tag = data.find('img')

            img_url = ""
            if img_tag:
                img_url = (
                    img_tag.attr('src') 
                    or img_tag.attr('data-src') 
                    or img_tag.attr('data-lazyload') 
                    or ""
                ).strip()

            if img_url.startswith("//"):
                img_url = "https:" + img_url
            elif img_url.startswith("/"):
                img_url = "https://cbu01.alicdn.com/" + img_url

            # 判断图片链接是否有效并下载
            if not img_url or not isinstance(img_url, str) or not img_url.startswith(('http', '//', '/')):
                self.logger.warning(f"[警告] 第 {self.count - 1} 个商品图片 URL 无效，跳过插图")
                self.logger.warning("商品 HTML 片段：")
                print(data.outer_html())
                image_path = None
            else:
                image_path = self.download_image(img_url, self.count - 1) if self.insert_image else None

            # 解析其他字段
            title = data.find('.offer-title-row .title-text div').text()

            price_container = data.find('.price-item')
            if price_container:
                # 用集合记录是否遇到 '¥'
                seen_yuan = set()  # 哈希表来记录是否遇到 '¥'
                price_parts = []
    
                for child in price_container.children():
                    part = pq(child).text().strip()
                    if part:  # 排除空白内容
                        # 如果遇到 '¥' 符号并且它已经出现过，停止继续添加
                        if '¥' in part:
                            if '¥' in seen_yuan:
                                break  # 如果 '¥' 已经出现过一次，停止处理
                            seen_yuan.add('¥')  # 记录第一次遇到 '¥'
                        price_parts.append(part)

                # 合并前两个价格部分
                price_text = ''.join(price_parts[:2]).replace('¥', '').strip()

                # 处理价格转换并控制小数点后两位
                try:
                    price = round(float(price_text), 2)
                except ValueError:
                    price = 0.0
            else:
                price = 0.0

            try:
                desc_text_elem = data.find('.col-desc_after .offer-desc-item .desc-text')
                deal_text = desc_text_elem.text().strip() if desc_text_elem else "0"

                if deal_text:
                    match = re.search(r'([\d\.]+)(万)?', deal_text)
                    if match:
                        number_part = match.group(1)
                        wan_unit = match.group(2)
                        deal = int(float(number_part) * (10000 if wan_unit else 1))
                    else:
                        deal = 0
                else:
                    deal = 0
            except Exception:
                deal = 0

            desc_text_elems = data.find('.col-desc').find('.desc-text')
            texts = [elem.text().strip() for elem in desc_text_elems.items()]
            post = "包邮" if any("包邮" in t for t in texts) else "/"

            try:
                shop_elem = data.find('.col-left a.offer-desc-item')
                shop = shop_elem.find('.desc-text').text().strip() if shop_elem else ''
                shop_url = shop_elem.attr('href').strip() if shop_elem else ''
                if shop_url.startswith("//"):
                    shop_url = "https:" + shop_url
            except Exception:
                shop = ''
                shop_url = ''

            item_url = data.attr('href')
            if item_url.startswith("//"):
                item_url = "https:" + item_url
            elif item_url.startswith("/"):
                item_url = "https://detail.1688.com" + item_url

            # 获取额外信息（地区、评论数）
            location, num_com = self.get_more(item_url)

            # 写入 Excel
            row = self.count
            self.sheet.row_dimensions[row].height = 65
            self.sheet.column_dimensions['L'].width = 13

            self.sheet.cell(row=row, column=1, value=row - 1)
            self.sheet.cell(row=row, column=2, value=title)
            self.sheet.cell(row=row, column=3, value=price)
            self.sheet.cell(row=row, column=4, value=deal)
            self.sheet.cell(row=row, column=5, value=location)
            self.sheet.cell(row=row, column=6, value=shop)
            self.sheet.cell(row=row, column=7, value=post)
            self.sheet.cell(row=row, column=8, value=item_url)
            self.sheet.cell(row=row, column=9, value=shop_url)
            self.sheet.cell(row=row, column=10, value=img_url)
            self.sheet.cell(row=row, column=11, value=num_com)

            # 插入图片
            if self.insert_image and image_path and os.path.exists(image_path):
                try:
                    img = ExcelImage(image_path)
                    img.width, img.height = 80, 80
                    self.sheet.add_image(img, f'L{row}')
                except Exception as e:
                    self.logger.error(f"插入图片失败: {e}（图片路径: {image_path}）")

            print(f"第{row - 1}个商品信息已保存")
            self.count += 1
            slide_counter += 1

            if slide_counter % 6 == 0:
                self.scroll_step_down()
            self.human_sleep(1, 2)

        return True


    def turn_page(self, page_number):
        try:
            self.go_to_page(page_number)
            # 等待页码跳转成功（可选：根据实际页码位置调整 XPath 或 CSS）
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.paging-info > em'),  # 修改为实际显示当前页码的位置
                str(page_number)
            ))
            self.check_for_captcha()
            self.human_sleep(2, 3)
        except Exception as e:
            print("翻页失败:", e)
            self.save_page_html("error_page.html")

    def run(self):
        self.driver.get("https://www.1688.com/")
        input("请在浏览器中手动登录1688，如出现滑块，请手动完成验证后按 Enter 继续......")
        self.search()
        # 切换到最新打开的标签页
        handles = self.driver.window_handles
        self.driver.switch_to.window(handles[-1])

        if self.page_start != 1:

            self.multi_scroll_up_down()
            self.go_to_page(self.page_start)

        for page in range(self.page_start, self.page_end + 1):
            if self.count - 2 >= self.max_items:  # 检查是否已达到最大抓取数量
                print(f"已抓取 {self.max_items} 个商品，停止抓取")
                break  # 达到最大数量时停止抓取

            self.multi_scroll_up_down(8,8,-800,800)
            proceed = self.parse_page(page)
            if not proceed:
                self.logger.warning(print(f"第 {page} 页爬取失败，跳过"))
                continue  # 如果该页抓取失败，跳过

            if page != self.page_end:
                self.turn_page(page + 1)

        filename = f"{self.keyword}_{time.strftime('%Y%m%d-%H%M')}.xlsx"
        self.excel.save(filename)
        print(f"文件已保存: {filename}")
        self.clean_image_folder()
        self.driver.quit()


if __name__ == "__main__":
    keyword = "奉贤黄桃"#input("输入搜索关键词：")
    start_page =1 # int(input("起始页码："))
    end_page = 3 #int(input("终止页码："))
    max_items = 50 #int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"

    scraper = Ali_1688Scraper(keyword, start_page, end_page, max_items=max_items)
    scraper.insert_image = insert_image
    scraper.run()

=======
﻿# -*- coding: utf-8 -*-

import time
import random
import json
import re
import requests
import os
import shutil

from io import BytesIO
from PIL import Image
from PIL import UnidentifiedImageError

from seleniumwire import webdriver  # selenium-wire 用于拦截请求
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from BaseScraper import BaseScraper



class Ali_1688Scraper(BaseScraper):
    def __init__(self, keyword, start_page, end_page, max_items=100):
        #super().__init__(headless=None, proxy=None)
        self.keyword = keyword
        self.page_start = start_page
        self.page_end = end_page
        self.max_items = max_items
        self.logger = self._init_logger()
        self.driver = self.init_driver()
        self.wait = WebDriverWait(self.driver, 10)
        self.excel = Workbook()
        self.sheet = self.excel.active
        self.count = 2
        self.insert_image = True
        self._setup_excel()
        #self.lock = threading.Lock()  # 用于同步的锁

    def human_sleep(self, min_time=0.5, max_time=1.5):
        time.sleep(random.uniform(min_time, max_time))

    def check_for_captcha(self):
        try:
            # 使用 self.driver 查找页面中特定的验证码元素
            captcha_element = self.driver.find_element_by_xpath("//div[@class='captcha']")
            if captcha_element:
                print("检测到验证码！暂停程序。")
                return True
        except Exception as e:
            # 如果没有找到验证码相关元素，返回 False
            return False




    def _setup_excel(self):
        headers = ['Num', 'Title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree',
                   'Title_URL', 'Shop_URL', 'Img_URL', 'Num_Com', 'Image']
        for i, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=i, value=header)

    def save_cookies(self, path="taobao_cookies.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.driver.get_cookies(), f)

    def load_cookies(self, path="taobao_cookies.json"):
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for cookie in cookies:
                self.driver.add_cookie(cookie)
        self.driver.refresh()

    def login(self, user, password):
        print("正在尝试登录1688...")
        try:
            iframe = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//iframe[contains(@src, "login.taobao.com")]')))
            self.driver.switch_to.frame(iframe)
            self.driver.find_element(By.ID, 'fm-login-id').send_keys(user)
            self.human_sleep(1, 2)
            self.driver.find_element(By.ID, 'fm-login-password').send_keys(password)
            self.human_sleep(1, 2)
            self.driver.find_element(By.CSS_SELECTOR, 'button.fm-button.fm-submit.password-login').click()
            self.human_sleep(1, 2)
            input("如出现滑块，请手动完成验证后按 Enter 继续...")
            self.save_cookies()
        except Exception as e:
            self.logger.waring(print("登录失败:", e))

    def search(self):
        try:
            print("尝试用定位搜索框和按钮...")
            search_box = self.wait.until(EC.element_to_be_clickable((By.ID, 'alisearch-input')))
            search_btn = self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'input-button')))
            self.check_for_captcha()
        except Exception as e:
            print("搜索框定位失败")
        try:
            search_box.clear()
            search_box.send_keys(self.keyword)
            self.human_sleep(1, 2)
            search_btn.click()
            self.check_for_captcha()
            self.human_sleep(2, 3)
        except Exception as e:
            self.logger.warning(print("搜索操作失败:", e))
            self.save_page_html("error_page.html")

    def go_to_page(self, page_number):
        try:
            # 定位输入框，清空并输入页码
            page_input = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'input.input-page')))
            page_input.clear()
            page_input.send_keys(str(page_number))
            self.human_sleep(0.5, 1.2)

            # 定位并点击“确定”按钮
            confirm_btn = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'div.paging-to-page-button')))
            confirm_btn.click()
            self.check_for_captcha()
            self.human_sleep(2, 3)

        except Exception as e:
            self.logger.warning(f"翻页失败: {e}")
            self.save_page_html("error_page.html")


    def get_more(self, url):
        comment_count = 0
        location = ''
        main_window = self.driver.current_window_handle

        try:
            # 提取商品ID，防止href匹配失败
            item_id_match = re.search(r'/offer/(\d+)\.html', url)
            if not item_id_match:
                self.logger.warning("无法提取商品ID，跳过")
                return location, comment_count
            item_id = item_id_match.group(1)

            # 限制 href 必须包含 detail.1688.com
            title_xpath = f'//a[contains(@href, "/offer/{item_id}.html") and contains(@href, "detail.1688.com")]'
            title_element = self.wait.until(EC.element_to_be_clickable((By.XPATH, title_xpath)))

            href = title_element.get_attribute('href')
            #print(f"即将点击链接：{href}")

            # 获取尺寸并设置更合适的偏移
            size = title_element.size
            width = size['width']
            height = size['height']
            offset_x = int(width * 0.1)
            offset_y = int(height * 0.3)

            # 执行偏移点击
            actions = ActionChains(self.driver)
            actions.key_down(Keys.CONTROL) \
                   .move_to_element_with_offset(title_element, offset_x, offset_y) \
                   .click() \
                   .key_up(Keys.CONTROL) \
                   .perform()

            self.check_for_captcha()
            # 等待新标签页打开，最多等待5秒
            for _ in range(10):
                handles = self.driver.window_handles
                if len(handles) > 1:
                    break
                time.sleep(0.5)
            else:
                self.logger.warning("新标签页未打开")
                self.save_page_html("error_page.html")
                return location, comment_count

            # 切换到新标签页
            handles = self.driver.window_handles
            new_tabs = [h for h in handles if h != main_window]
            if not new_tabs:
                self.logger.warning("找不到新标签页句柄")
                self.save_page_html("error_page.html")
                return location, comment_count

            new_tab = new_tabs[-1]
            self.driver.switch_to.window(new_tab)
            self.check_for_captcha()
            self.human_sleep(2, 3)

            # 获取评论数
            try:
                element = self.driver.find_element(By.CSS_SELECTOR, 'span.brackets')
                comment_count = int(element.get_attribute("data-value"))
            except Exception as e:
                self.logger.warning(f"获取评论数失败，错误信息：{e}")
                comment_count = 0

            # 获取发货地
            try:
                location_element = self.driver.find_element(By.CSS_SELECTOR, 'span.location')
                location = location_element.text.strip()
            except Exception as e:
                self.logger.warning(f"获取发货地失败，错误信息：{e}")
                location = ""

        except Exception as e:
            self.logger.warning(f"标签页打开失败，错误信息：{e}")

        finally:
            # 关闭新标签页，切换回主窗口
            self.human_sleep(1, 2)
            try:
                current_handle = self.driver.current_window_handle
                if current_handle != main_window:
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    self.check_for_captcha()

            except Exception as e:
                self.logger.warning(f"关闭标签页或切换主窗口失败，错误信息：{e}")
                self.save_page_html("error_page.html")

        return location, comment_count

    def download_image(self, url, index):
        try:
            # 补全 url
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/"):
                url = "https://img.alicdn.com" + url

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer": "https://www.1688.com/",
            }

            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                self.logger.error(print(f"图片请求失败: {url}"))
                return None

            try:
                # 读取图片并转成 RGB
                img = Image.open(BytesIO(response.content)).convert("RGB")
            except UnidentifiedImageError:
                self.logger.error(print(f"图片解码失败（格式可能不受支持，如webp）: {url}"))
                return None
            except Exception as e:
                self.logger.error(print(f"图片处理异常: {e} | URL: {url}"))
                return None

            # 创建文件夹
            folder = "images"
            if not os.path.exists(folder):
                os.makedirs(folder)

            # 生成文件名，保存为 jpg
            file_name = f"{index:04d}.jpg"
            file_path = os.path.join(folder, file_name)

            # 保存成 jpg 文件
            img.save(file_path, format="JPEG")

            # 返回文件路径
            return file_path

        except Exception as e:
            self.logger.error(print(f"图片下载异常: {e} | URL: {url}"))
            return None

    def clean_image_folder(self):
        folder = "images"
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)  # 删除文件或快捷方式
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)  # 删除子文件夹

    ##滑动窗口
    def scroll_step_down(self, base_step=300):
        """模拟人类向下小段滑动"""
        step = random.randint(base_step - 100, base_step + 100)
        self.driver.execute_script(f"window.scrollBy(0, {step});")
        time.sleep(random.uniform(0.6, 1.2))
        #self.wait_manual_verification()


    def smart_scroll_until_loaded(self, min_new_items=4, max_scroll_attempts=10):
        """边滚动边检测新商品是否加载，最多尝试 max_scroll_attempts 次"""
        last_count = 0
        attempts = 0

    def multi_scroll_up_down(self, up_num=1,down_num=8 , up_step=-200, down_step=600):
        # 向下滚动
        for _ in range(down_num):
            self.scroll_step_down(down_step)
    
        # 向上滚动
        for _ in range(up_num):
            self.scroll_step_down(up_step)
 

        # 模拟人工休眠，防止过快的滚动
        time.sleep(random.uniform(1.0, 1.5))

    def parse_page(self, page_number):
        # 获取当前页面所有商品元素（Selenium WebElement）
        items_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a.search-offer-item')

        # 用 pyquery 解析页面HTML，拿到所有商品节点
        html = self.driver.page_source
        doc = pq(html)
        items_data = list(doc('a.search-offer-item').items())

        if len(items_elements) != len(items_data):
            self.logger.warning(f"警告：Selenium 找到商品数量 {len(items_elements)} 与 pyquery 找到数量 {len(items_data)} 不一致，可能页面未完全加载！")
            # 这里可以等待或重试，这里先简单返回False停止
            return False

        slide_counter = 0  # 下滑计数器

        for idx, (element, data) in enumerate(zip(items_elements, items_data)):
            if self.count - 2 >= self.max_items:
                print(f"已达到最大抓取数量：{self.max_items}，停止抓取。")
                return False

            # 滚动到当前商品，使其位于中央
            self.driver.execute_script("""
                var elem = arguments[0];
                var rect = elem.getBoundingClientRect();
                var windowHeight = window.innerHeight;
                var elementTop = rect.top;
                var elementBottom = rect.bottom;
                var scrollAmount = (elementTop + elementBottom) / 2 - windowHeight / 2;
                window.scrollBy(0, scrollAmount);
            """, element)

            # 解析图片URL
            img_tag = data.find('.offer-img-inner img') or data.find('img')
            if not img_tag:
                img_tag = data.find('img')

            img_url = ""
            if img_tag:
                img_url = (
                    img_tag.attr('src') 
                    or img_tag.attr('data-src') 
                    or img_tag.attr('data-lazyload') 
                    or ""
                ).strip()

            if img_url.startswith("//"):
                img_url = "https:" + img_url
            elif img_url.startswith("/"):
                img_url = "https://cbu01.alicdn.com/" + img_url

            # 判断图片链接是否有效并下载
            if not img_url or not isinstance(img_url, str) or not img_url.startswith(('http', '//', '/')):
                self.logger.warning(f"[警告] 第 {self.count - 1} 个商品图片 URL 无效，跳过插图")
                self.logger.warning("商品 HTML 片段：")
                print(data.outer_html())
                image_path = None
            else:
                image_path = self.download_image(img_url, self.count - 1) if self.insert_image else None

            # 解析其他字段
            title = data.find('.offer-title-row .title-text div').text()

            price_container = data.find('.price-item')
            if price_container:
                # 用集合记录是否遇到 '¥'
                seen_yuan = set()  # 哈希表来记录是否遇到 '¥'
                price_parts = []
    
                for child in price_container.children():
                    part = pq(child).text().strip()
                    if part:  # 排除空白内容
                        # 如果遇到 '¥' 符号并且它已经出现过，停止继续添加
                        if '¥' in part:
                            if '¥' in seen_yuan:
                                break  # 如果 '¥' 已经出现过一次，停止处理
                            seen_yuan.add('¥')  # 记录第一次遇到 '¥'
                        price_parts.append(part)

                # 合并前两个价格部分
                price_text = ''.join(price_parts[:2]).replace('¥', '').strip()

                # 处理价格转换并控制小数点后两位
                try:
                    price = round(float(price_text), 2)
                except ValueError:
                    price = 0.0
            else:
                price = 0.0

            try:
                desc_text_elem = data.find('.col-desc_after .offer-desc-item .desc-text')
                deal_text = desc_text_elem.text().strip() if desc_text_elem else "0"

                if deal_text:
                    match = re.search(r'([\d\.]+)(万)?', deal_text)
                    if match:
                        number_part = match.group(1)
                        wan_unit = match.group(2)
                        deal = int(float(number_part) * (10000 if wan_unit else 1))
                    else:
                        deal = 0
                else:
                    deal = 0
            except Exception:
                deal = 0

            desc_text_elems = data.find('.col-desc').find('.desc-text')
            texts = [elem.text().strip() for elem in desc_text_elems.items()]
            post = "包邮" if any("包邮" in t for t in texts) else "/"

            try:
                shop_elem = data.find('.col-left a.offer-desc-item')
                shop = shop_elem.find('.desc-text').text().strip() if shop_elem else ''
                shop_url = shop_elem.attr('href').strip() if shop_elem else ''
                if shop_url.startswith("//"):
                    shop_url = "https:" + shop_url
            except Exception:
                shop = ''
                shop_url = ''

            item_url = data.attr('href')
            if item_url.startswith("//"):
                item_url = "https:" + item_url
            elif item_url.startswith("/"):
                item_url = "https://detail.1688.com" + item_url

            # 获取额外信息（地区、评论数）
            location, num_com = self.get_more(item_url)

            # 写入 Excel
            row = self.count
            self.sheet.row_dimensions[row].height = 65
            self.sheet.column_dimensions['L'].width = 13

            self.sheet.cell(row=row, column=1, value=row - 1)
            self.sheet.cell(row=row, column=2, value=title)
            self.sheet.cell(row=row, column=3, value=price)
            self.sheet.cell(row=row, column=4, value=deal)
            self.sheet.cell(row=row, column=5, value=location)
            self.sheet.cell(row=row, column=6, value=shop)
            self.sheet.cell(row=row, column=7, value=post)
            self.sheet.cell(row=row, column=8, value=item_url)
            self.sheet.cell(row=row, column=9, value=shop_url)
            self.sheet.cell(row=row, column=10, value=img_url)
            self.sheet.cell(row=row, column=11, value=num_com)

            # 插入图片
            if self.insert_image and image_path and os.path.exists(image_path):
                try:
                    img = ExcelImage(image_path)
                    img.width, img.height = 80, 80
                    self.sheet.add_image(img, f'L{row}')
                except Exception as e:
                    self.logger.error(f"插入图片失败: {e}（图片路径: {image_path}）")

            print(f"第{row - 1}个商品信息已保存")
            self.count += 1
            slide_counter += 1

            if slide_counter % 6 == 0:
                self.scroll_step_down()
            self.human_sleep(1, 2)

        return True


    def turn_page(self, page_number):
        try:
            self.go_to_page(page_number)
            # 等待页码跳转成功（可选：根据实际页码位置调整 XPath 或 CSS）
            self.wait.until(EC.text_to_be_present_in_element(
                (By.CSS_SELECTOR, 'span.paging-info > em'),  # 修改为实际显示当前页码的位置
                str(page_number)
            ))
            self.check_for_captcha()
            self.human_sleep(2, 3)
        except Exception as e:
            print("翻页失败:", e)
            self.save_page_html("error_page.html")

    def run(self):
        self.driver.get("https://www.1688.com/")
        input("请在浏览器中手动登录1688，如出现滑块，请手动完成验证后按 Enter 继续......")
        self.search()
        # 切换到最新打开的标签页
        handles = self.driver.window_handles
        self.driver.switch_to.window(handles[-1])

        if self.page_start != 1:

            self.multi_scroll_up_down()
            self.go_to_page(self.page_start)

        for page in range(self.page_start, self.page_end + 1):
            if self.count - 2 >= self.max_items:  # 检查是否已达到最大抓取数量
                print(f"已抓取 {self.max_items} 个商品，停止抓取")
                break  # 达到最大数量时停止抓取

            self.multi_scroll_up_down(8,8,-800,800)
            proceed = self.parse_page(page)
            if not proceed:
                self.logger.warning(print(f"第 {page} 页爬取失败，跳过"))
                continue  # 如果该页抓取失败，跳过

            if page != self.page_end:
                self.turn_page(page + 1)

        filename = f"{self.keyword}_{time.strftime('%Y%m%d-%H%M')}.xlsx"
        self.excel.save(filename)
        print(f"文件已保存: {filename}")
        self.clean_image_folder()
        self.driver.quit()


if __name__ == "__main__":
    keyword = "奉贤黄桃"#input("输入搜索关键词：")
    start_page =1 # int(input("起始页码："))
    end_page = 3 #int(input("终止页码："))
    max_items = 50 #int(input("最多抓取商品数量："))
    insert_image = input("是否插入商品图片到 Excel？(y/n)：").strip().lower() == "y"

    scraper = Ali_1688Scraper(keyword, start_page, end_page, max_items=max_items)
    scraper.insert_image = insert_image
    scraper.run()

>>>>>>> 6727600f6c5eaac5434322fc753ddbbebeda62da
